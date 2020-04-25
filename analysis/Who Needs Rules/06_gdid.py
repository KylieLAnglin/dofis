#!/usr/bin/env python
# coding: utf-8

# %%


import os
import sys

import matplotlib.mlab as mlab
import matplotlib.pyplot as plt
import numpy as np
import pandas as pd
import statsmodels.formula.api as smf
from linearmodels import PanelOLS
from openpyxl import load_workbook
from patsy import dmatrices

sys.path.append("../")
from library import start, analysis


# %%

data = pd.read_csv(os.path.join(start.data_path, 'clean', 'gdid_subject.csv'),
                   sep=",", low_memory=False)
print(data[(data.doi)].district.nunique())
data = data[(data.doi)]
print(data.doi_year.value_counts())
data.sample()

# %%

# convert year to datetime
df = data.reset_index()
df['year'] = pd.to_datetime(df['year'], format='%Y')
# add column year to index
df = data.set_index(['year', 'campus'])
# swap indexes
df.index = df.index.swaplevel(0, 1)
df[['district', 'doi_year', 'treatpost']].sample(5)

# %% Get table ready

gdid_model = 'score_std ~ + 1 + treatpost + C(test_by_year) + EntityEffects'
linear_gdid_model = 'score_std ~ + 1 + treatpost + yearpost + ' \
    'yearpre  + C(test_by_year) + EntityEffects'
event_study_model = 'score_std ~ + 1 + pre5 + pre4 + pre3 + pre2 + ' \
    'post1 + post2 + post3  + C(test_by_year) + EntityEffects'

file_name = start.table_path + 'table3_gdid_and_event_math.xlsx'
wb = load_workbook(file_name)
ws = wb.active

# %%


def results_table(data, file_name):
    file = start.table_path + file_name
    wb = load_workbook(file)
    ws = wb.active

    # GDID
    mod = PanelOLS.from_formula(gdid_model, data)
    res = mod.fit(cov_type='clustered', clusters=data.district)
    # res = mod.fit(cov_type='clustered', cluster_entity = True,
    # cluster_time = True)
    print(res)
    ws.cell(row=3, column=2).value = analysis.coef_with_stars(
        res.params['treatpost[T.True]'], res.pvalues['treatpost[T.True]'])
    ws.cell(row=4, column=2).value = analysis.format_se(
        res.std_errors['treatpost[T.True]'])

    # Linear GDID
    mod = PanelOLS.from_formula(linear_gdid_model, data)
    res = mod.fit(cov_type='clustered', clusters=data.district)
    print(res)
    ws.cell(row=6, column=2).value = analysis.coef_with_stars(
        res.params['treatpost[T.True]'], res.pvalues['treatpost[T.True]'])
    ws.cell(row=7, column=2).value = analysis.format_se(
        res.std_errors['treatpost[T.True]'])
    ws.cell(row=8, column=2).value = analysis.coef_with_stars(
        res.params['yearpost'], res.pvalues['yearpost'])
    ws.cell(row=9, column=2).value = analysis.format_se(
        res.std_errors['yearpost'])
    ws.cell(row=10, column=2).value = analysis.coef_with_stars(
        res.params['yearpre'], res.pvalues['yearpre'])
    ws.cell(row=11, column=2).value = analysis.format_se(
        res.std_errors['yearpre'])

    # Event Study
    mod = PanelOLS.from_formula(event_study_model, data)
    res = mod.fit(cov_type='clustered', clusters=data.district)
    print(res)
    row = 3
    for coef in ['post3', 'post2', 'post1', 'pre2', 'pre3', 'pre4', 'pre5']:
        ws.cell(row=row, column=4).value = analysis.coef_with_stars(
            res.params[coef], res.pvalues[coef])
        row = row + 1
        ws.cell(row=row, column=4).value = analysis.format_se(
            res.std_errors[coef])
        row = row + 1

    wb.save(file)

# %% Math


results_table(df[df.math == 1], 'table3_gdid_and_event_math.xlsx')

results_table(df[df.reading == 1], 'table4_gdid_and_event_reading.xlsx')

# %% Event Study Graphs - Math

mod = PanelOLS.from_formula(event_study_model, df[df.math == 1])
res = mod.fit(cov_type='clustered', clusters=df[df.math == 1].district)
nonparametric = []
nonparametric_se = []
for coef in ['pre5', 'pre4', 'pre3', 'pre2', 'pre1',
             'post1', 'post2', 'post3']:
    nonpar = 0
    nonpar_se = 0
    if coef != 'pre1':
        nonpar = res.params[coef]
        nonpar_se = res.std_errors[coef]
    nonparametric.append(nonpar)
    nonparametric_se.append(nonpar_se)
coef_df = pd.DataFrame({'coef': nonparametric,
                        'err': nonparametric_se,
                        'year': [-5, -4, -3, -2, -1, 1, 2, 3]
                        })
coef_df['lb'] = coef_df.coef - (1.96*coef_df.err)
coef_df['ub'] = coef_df.coef + (1.96*coef_df.err)
coef_df['errsig'] = coef_df.err * 1.96

fig, ax = plt.subplots(figsize=(8, 5))

coef_df.plot(x='year', y='coef', kind='bar',
             ax=ax, color='none',
             yerr='errsig', legend=False)
ax.set_ylabel('')
ax.set_xlabel('')
ax.scatter(x=pd.np.arange(coef_df.shape[0]),
           marker='s', s=120,
           y=coef_df['coef'], color='black')
ax.axhline(y=0, linestyle='--', color='black', linewidth=4)
ax.xaxis.set_ticks_position('none')
_ = ax.set_xticklabels(['Pre5', 'Pre4', 'Pre3', 'Pre2', 'Pre1',
                        'Post1', 'Post2', 'Post3'],
                       rotation=0)
# ax.set_title('Impact on Student Achievement - Event Study Coefficients',
# fontsize = 16)

fig.savefig(start.table_path + 'math_event_study' + '.png',
            bbox_inches="tight")


# %% Event Study Graphs - Reading

mod = PanelOLS.from_formula(event_study_model, df[df.reading == 1])
res = mod.fit(cov_type='clustered', clusters=df[df.reading == 1].district)
nonparametric = []
nonparametric_se = []
for coef in ['pre5', 'pre4', 'pre3', 'pre2',
             'pre1', 'post1', 'post2', 'post3']:
    nonpar = 0
    nonpar_se = 0
    if coef != 'pre1':
        nonpar = res.params[coef]
        nonpar_se = res.std_errors[coef]
    nonparametric.append(nonpar)
    nonparametric_se.append(nonpar_se)
coef_df = pd.DataFrame({'coef': nonparametric,
                        'err': nonparametric_se,
                        'year': [-5, -4, -3, -2, -1, 1, 2, 3]
                        })
coef_df['lb'] = coef_df.coef - (1.96*coef_df.err)
coef_df['ub'] = coef_df.coef + (1.96*coef_df.err)
coef_df['errsig'] = coef_df.err * 1.96

fig, ax = plt.subplots(figsize=(8, 5))

coef_df.plot(x='year', y='coef', kind='bar',
             ax=ax, color='none',
             yerr='errsig', legend=False)
ax.set_ylabel('')
ax.set_xlabel('')
ax.scatter(x=pd.np.arange(coef_df.shape[0]),
           marker='s', s=120,
           y=coef_df['coef'], color='black')
ax.axhline(y=0, linestyle='--', color='black', linewidth=4)
ax.xaxis.set_ticks_position('none')
_ = ax.set_xticklabels(['Pre5', 'Pre4', 'Pre3', 'Pre2', 'Pre1',
                        'Post1', 'Post2', 'Post3'],
                       rotation=0)
# ax.set_title('Impact on Student Achievement - Event Study Coefficients',
#  fontsize = 16)

fig.savefig(start.table_path + 'reading_event_study' + '.png',
            bbox_inches="tight")


# %% Table by Subject
subjects = ['m_3rd_avescore', 'm_4th_avescore',  'm_5th_avescore',
            'm_6th_avescore', 'm_7th_avescore', 'm_8th_avescore',
            'alg_avescore',
            'r_3rd_avescore', 'r_4th_avescore', 'r_5th_avescore',
            'r_6th_avescore', 'r_7th_avescore', 'r_8th_avescore',
            'eng1_avescore',  'bio_avescore']


# All Subject Table
file = start.table_path + 'tableA_effect_by_subject.xlsx'
wb = load_workbook(file)
ws = wb.active

col = 3
for subject in subjects:

    df_sub = df[df.test == subject]
    test = pd.Categorical(df_sub.test)

    # GDID
    mod = PanelOLS.from_formula(gdid_model, df_sub)
    res = mod.fit(cov_type='clustered', clusters=df_sub.district)
    ws.cell(row=5, column=col).value = analysis.bonferroni(
        len(subjects),
        res.params['treatpost[T.True]'],
        res.pvalues['treatpost[T.True]'])
    ws.cell(row=6, column=col).value = analysis.format_se(
        res.std_errors['treatpost[T.True]'])

    # GDID with Trend
    mod = PanelOLS.from_formula(linear_gdid_model, df_sub)
    res = mod.fit(cov_type='clustered', clusters=df_sub.district)
    jump = res.params['treatpost[T.True]']
    preslope = res.params['yearpre']
    preslope_se = res.std_errors['yearpre']
    postslope = res.params['yearpost']
    post_slope = res.std_errors['yearpost']
    ws.cell(row=8, column=col).value = analysis.bonferroni(
        len(subjects),
        res.params['treatpost[T.True]'],
        res.pvalues['treatpost[T.True]'])
    ws.cell(row=9, column=col).value = analysis.format_se(
        res.std_errors['treatpost[T.True]'])
    ws.cell(row=10, column=col).value = analysis.bonferroni(
        len(subjects), res.params['yearpost'], res.pvalues['yearpost'])
    ws.cell(row=11, column=col).value = analysis.format_se(
        res.std_errors['yearpost'])
    ws.cell(row=12, column=col).value = analysis.bonferroni(
        len(subjects), res.params['yearpre'], res.pvalues['yearpre'])
    ws.cell(row=13, column=col).value = analysis.format_se(
        res.std_errors['yearpre'])
    wb.save(file)

    # Event Study
    mod = PanelOLS.from_formula(event_study_model, df_sub)
    res = mod.fit(cov_type='clustered', clusters=df_sub.district)
    nonparametric = []
    nonparametric_se = []
    for coef in ['pre5', 'pre4', 'pre3', 'pre2', 'pre1',
                 'post1', 'post2', 'post3']:
        nonpar = 0
        nonpar_se = 0
        if coef != 'pre1':
            nonpar = res.params[coef]
            nonpar_se = res.std_errors[coef]
        nonparametric.append(nonpar)
        nonparametric_se.append(nonpar_se)
    row = 15
    for coef in ['post3', 'post2', 'post1', 'pre2', 'pre3', 'pre4', 'pre5']:
        ws.cell(row=row, column=col).value = analysis.coef_with_stars(
            res.params[coef], res.pvalues[coef])
        row = row + 1
        ws.cell(row=row, column=col).value = analysis.format_se(
            res.std_errors[coef])
        row = row + 1
    col = col + 1
wb.save(file)
