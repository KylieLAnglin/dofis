
# %%

import os
import sys

import numpy as np
import pandas as pd
import statsmodels.formula.api as smf
from linearmodels import PanelOLS
from openpyxl import load_workbook
from patsy import dmatrices

sys.path.append("../")
from library import start, analysis


# %%

data = pd.read_csv(os.path.join(start.data_path, 'clean', 'gdid_subject.csv'),
                   sep=",", low_memory=False)
data = data[(data.doi)]
print(data[(data.doi)].district.nunique())
print(data.doi_year.value_counts())
data.sample()

# %%

# convert year to datetime
df = data.reset_index()
df['year'] = pd.to_datetime(df['year'], format='%Y')
# add column year to index
df = data.set_index(['year', 'campus'])
# swap indexes
df.index = df.index.swaplevel(0, 1)
df[['district', 'doi_year', 'treatpost']].sample(5)

# %% Table by Subject
subjects = ['m_3rd_avescore', 'm_4th_avescore',  'm_5th_avescore',
            'm_6th_avescore', 'm_7th_avescore', 'm_8th_avescore',
            'alg_avescore',
            'r_3rd_avescore', 'r_4th_avescore', 'r_5th_avescore',
            'r_6th_avescore', 'r_7th_avescore', 'r_8th_avescore',
            'eng1_avescore',  'bio_avescore']

gdid_model = 'score_std ~ + 1 + treatpost + C(test_by_year) + EntityEffects'
linear_gdid_model = 'score_std ~ + 1 + treatpost + yearpost + ' \
    'yearpre  + C(test_by_year) + EntityEffects'
event_study_model = 'score_std ~ + 1 + pre5 + pre4 + pre3 + pre2 + ' \
    'post1 + post2 + post3  + C(test_by_year) + EntityEffects' 
    
# All Subject Table
file = start.table_path + 'tableA_effect_by_subject.xlsx'
wb = load_workbook(file)
ws = wb.active

col = 3
for subject in subjects:

    df_sub = df[df.test == subject]
    test = pd.Categorical(df_sub.test)

    # GDID
    mod = PanelOLS.from_formula(gdid_model, df_sub)
    res = mod.fit(cov_type='clustered', clusters=df_sub.district)
    ws.cell(row=5, column=col).value = analysis.bonferroni(
        len(subjects),
        res.params['treatpost[T.True]'],
        res.pvalues['treatpost[T.True]'])
    ws.cell(row=6, column=col).value = analysis.format_se(
        res.std_errors['treatpost[T.True]'])

    # GDID with Trend
    mod = PanelOLS.from_formula(linear_gdid_model, df_sub)
    res = mod.fit(cov_type='clustered', clusters=df_sub.district)
    jump = res.params['treatpost[T.True]']
    preslope = res.params['yearpre']
    preslope_se = res.std_errors['yearpre']
    postslope = res.params['yearpost']
    post_slope = res.std_errors['yearpost']
    ws.cell(row=8, column=col).value = analysis.bonferroni(
        len(subjects),
        res.params['treatpost[T.True]'],
        res.pvalues['treatpost[T.True]'])
    ws.cell(row=9, column=col).value = analysis.format_se(
        res.std_errors['treatpost[T.True]'])
    ws.cell(row=10, column=col).value = analysis.bonferroni(
        len(subjects), res.params['yearpost'], res.pvalues['yearpost'])
    ws.cell(row=11, column=col).value = analysis.format_se(
        res.std_errors['yearpost'])
    ws.cell(row=12, column=col).value = analysis.bonferroni(
        len(subjects), res.params['yearpre'], res.pvalues['yearpre'])
    ws.cell(row=13, column=col).value = analysis.format_se(
        res.std_errors['yearpre'])
    wb.save(file)

    # Event Study
    mod = PanelOLS.from_formula(event_study_model, df_sub)
    res = mod.fit(cov_type='clustered', clusters=df_sub.district)
    nonparametric = []
    nonparametric_se = []
    for coef in ['pre5', 'pre4', 'pre3', 'pre2', 'pre1',
                 'post1', 'post2', 'post3']:
        nonpar = 0
        nonpar_se = 0
        if coef != 'pre1':
            nonpar = res.params[coef]
            nonpar_se = res.std_errors[coef]
        nonparametric.append(nonpar)
        nonparametric_se.append(nonpar_se)
    row = 15
    for coef in ['post3', 'post2', 'post1', 'pre2', 'pre3', 'pre4', 'pre5']:
        ws.cell(row=row, column=col).value = analysis.coef_with_stars(
            res.params[coef], res.pvalues[coef])
        row = row + 1
        ws.cell(row=row, column=col).value = analysis.format_se(
            res.std_errors[coef])
        row = row + 1
        
    ws.cell(row=29, column=col).value = len(df_sub)
    ws.cell(row=30, column=col).value = df_sub.reset_index().campus.nunique()
    ws.cell(row=31, column=col).value = df_sub.district.nunique()

    col = col + 1

 
wb.save(file)


# %%
