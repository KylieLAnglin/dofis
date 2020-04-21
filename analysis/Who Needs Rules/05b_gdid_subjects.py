#!/usr/bin/env python
# coding: utf-8

# In[1]:


import os
import sys

import matplotlib.mlab as mlab
import matplotlib.pyplot as plt
import numpy as np
import pandas as pd
import statsmodels.formula.api as smf
from linearmodels import PanelOLS
from openpyxl import load_workbook
from patsy import dmatrices

sys.path.append("../")
from library import start, analysis


# In[5]:

data = pd.read_csv(os.path.join(start.data_path, 'clean', 'gdid_subject.csv'),
                  sep=",", low_memory= False)
print(data[(data.doi)].district.nunique())
data = data[(data.doi)]
print(data.doi_year.value_counts())
data.sample()

# In[12]:

# convert year to datetime
df = data.reset_index()
df['year'] = pd.to_datetime(df['year'], format='%Y')
# add column year to index
df = data.set_index(['year', 'campus'])
# swap indexes
df.index = df.index.swaplevel(0,1)
df[['district', 'doi_year','treatpost']].sample(5)



# %% Get table ready
file_name = start.table_path + 'table3_gdid_and_event_math.xlsx'
wb = load_workbook(file_name)
ws = wb.active


# %% Simple GDID

test = pd.Categorical(df.test)
mod = PanelOLS.from_formula(
    "avescores ~ + 1 + treatpost + C(test) + TimeEffects + EntityEffects'", df)
res = mod.fit(cov_type='clustered', cluster_entity=True)
print(res)
ws.cell(row=3, column=2).value = analysis.coef_with_stars(
    res.params['treatpost[T.True]'],
    res.pvalues['treatpost[T.True]'])
ws.cell(row=4, column=2).value = analysis.format_se(
    res.std_errors['treatpost[T.True]'])


# ## GDID with Trends

# In[15]:


mod = PanelOLS.from_formula('avescores ~ + 1 + treatpost + yearpost + yearpre + students_hisp + students_num + C(test) + TimeEffects + EntityEffects', df)
res = mod.fit(cov_type='clustered', cluster_entity=True)
print(res)
#intercept = res.params['Intercept']
jump = res.params['treatpost[T.True]']
preslope = res.params['yearpre']
postslope = res.params['yearpost']
ws.cell(row= 6, column= 2).value = coef_with_stars(res.params['treatpost[T.True]'], res.pvalues['treatpost[T.True]'])
ws.cell(row= 7, column= 2).value = format_se(res.std_errors['treatpost[T.True]'])
ws.cell(row= 8, column= 2).value = coef_with_stars(res.params['yearpost'], res.pvalues['treatpost[T.True]'])
ws.cell(row= 9, column= 2).value = format_se(res.std_errors['yearpost'])
ws.cell(row= 10, column= 2).value = coef_with_stars(res.params['yearpre'], res.pvalues['treatpost[T.True]'])
ws.cell(row= 11, column= 2).value = format_se(res.std_errors['yearpre'])
wb.save(file)


# In[16]:


mod = PanelOLS.from_formula('avescores ~ + 1 + pre5 + pre4 + pre3 + pre2 + pre1 + post1 + post2 + post3 + students_hisp + students_num + C(test) + TimeEffects + EntityEffects', df)
res = mod.fit(cov_type='clustered', cluster_entity=True)
print(res)
nonparametric = []
for coef in ['pre5', 'pre4', 'pre3', 'pre2', 'pre1', 'pre0', 'post1', 'post2', 'post3']:
    nonpar = 0
    if coef != 'pre0':
        nonpar = res.params[coef]
    nonparametric.append(nonpar)
print(nonparametric)
row = 3
for coef in ['post3', 'post2', 'post1', 'pre1', 'pre2', 'pre3', 'pre4', 'pre5']:
    ws.cell(row= row, column= 4).value = coef_with_stars(res.params[coef], res.pvalues[coef])
    row = row + 1
    ws.cell(row= row, column= 4).value = format_se(res.std_errors[coef])  
    row = row + 1
wb.save(file)


# # Create graph (look up how to use predicted values)
# 

# In[79]:


years = [-5, -4, -3, -2, -1, 0, 1, 2, 3]
parametric = []
for year in years:
    par = 0
    if year < 0 :
        par = (year * preslope)
    if year > 0 :
        par = jump + (year * postslope)
    parametric.append(par)
parametric   


# In[80]:


nonparametric


# In[81]:


plt.plot(years, parametric)
plt.plot(years, nonparametric)


# # Table

# In[ ]:





# # Table by Subject

# In[26]:


file = table_path + 'table3_gdid.xlsx'
row_n = 3
col_n = 2
wb = load_workbook(file)
ws = wb.active
outcomes = ['elem_math', 'elem_reading', 
            'middle_math', 'middle_reading', 'middle_science', 
            'biology', 'algebra', 'eng1']
for outcome in outcomes:
    model = outcome + ' ~ + 1 + treatpost + TimeEffects + EntityEffects'
    mod = PanelOLS.from_formula(model, df)
    res = mod.fit(cov_type='clustered', cluster_entity=True)
    p = res.pvalues['treatpost[T.True]']
    coef = res.params['treatpost[T.True]'].round(2)
    se = res.std_errors['treatpost[T.True]'].round(2)
    if p >= .05:
        coef = str(coef)
    if p < .05 and p > .01:
        coef = str(coef) + '*'
    if p < .01 and p > .001:
        coef = str(coef) + '**'
    if p < .001:
        coef = str(coef) + '***'
    ws.cell(row=row_n, column=col_n).value = coef
    row_n = row_n + 1
    ws.cell(row=row_n, column=col_n).value = '(' + str(se) + ')'
    row_n = row_n + 1
wb.save(file)


# In[27]:


file = table_path + 'table3_gdid.xlsx'
row_n = 3
col_n = 3
wb = load_workbook(file)
ws = wb.active
outcomes = ['elem_math', 'elem_reading', 
            'middle_math', 'middle_reading', 'middle_science', 
            'biology', 'algebra', 'eng1']
for outcome in outcomes:
    model = outcome + ' ~ + 1 + treatpost + students_hisp + students_num + TimeEffects + EntityEffects'
    mod = PanelOLS.from_formula(model, df)
    res = mod.fit(cov_type='clustered', cluster_entity=True)
    print(res)
    p = res.pvalues['treatpost[T.True]']
    coef = res.params['treatpost[T.True]'].round(2)
    se = res.std_errors['treatpost[T.True]'].round(2)
    print(p)
    if p >= .05:
        coef = str(coef)
    if p < .05 and p > .01:
        coef = str(coef) + '*'
    if p < .01 and p > .001:
        coef = str(coef) + '**'
    if p < .001:
        coef = str(coef) + '***'
    ws.cell(row=row_n, column=col_n).value = coef
    row_n = row_n + 1
    ws.cell(row=row_n, column=col_n).value = '(' + str(se) + ')'
    row_n = row_n + 1
wb.save(file)


# In[ ]:


model = outcome + ' ~ + 1 + treatpost + students_hisp + students_num + TimeEffects + EntityEffects'
    mod = PanelOLS.from_formula(model, df)
    res = mod.fit(cov_type='clustered', cluster_entity=True)


# # Heterogenous effects

# In[7]:


### Generate variable for number of hispanic students in year before the district declares


# In[8]:


data.teachers_turnover_ratio_d


# In[9]:


data_pre = data.loc[data.yearpost == -1]
data_pre = data_pre.rename(columns = {'students_hisp': 'students_hisp_pre',
                                     'teachers_turnover_ratio_d': 'teacher_turnover_pre'})
data_pre['low_avescores_pre'] = np.where(data_pre.avescores < data_pre.avescores.quantile(.25),1, 0)
data_pre['high_avescores_pre'] = np.where(data_pre.avescores > data_pre.avescores.quantile(.75),1,0)
data_pre = data_pre[['campus','students_hisp_pre', 'teacher_turnover_pre', 'low_avescores_pre', 'high_avescores_pre']]
data_hte = data.merge(data_pre, on = 'campus', how = 'left')
data_hte['majority_hisp_pre'] = np.where(data_hte.students_hisp_pre > .6, 1, 0)
data_hte['high_turnover'] = np.where(data_hte.teacher_turnover_pre > 20, 1, 0)
data_hte


# In[10]:


data_pre = data.loc[data.year == 2016]
data_pre['rural_pre'] = np.where(data_pre.type_description == "RURAL", 1, 0)
data_pre['urban_pre'] = np.where(data_pre.type_description == "URBAN", 1, 0)
data_pre = data_pre[['campus','rural_pre', 'urban_pre']]
data_hte = data_hte.merge(data_pre, on = 'campus', how = 'left')
#df_hte[['students_hisp', 'students_hisp_pre', 'yearpost']].sample(5)


# In[11]:


#convert year to datetime
df_hte = data_hte.reset_index()
df_hte['year'] = pd.to_datetime(df_hte['year'], format='%Y')
#add column year to index
df_hte = data_hte.set_index(['year', 'campus'])
#swap indexes
df_hte.index = df_hte.index.swaplevel(0,1)
df_hte[['district', 'doi_year','treatpost',]].tail(20)


# # Majority Hispanic

# In[12]:


df_hte['treatpost_hisp'] = df_hte.treatpost * df_hte.majority_hisp_pre
mod = PanelOLS.from_formula('avescores ~ + 1 + treatpost + treatpost_hisp + students_hisp + students_num + TimeEffects + EntityEffects', df_hte)
res = mod.fit(cov_type='clustered', cluster_entity=True)
print(res)


# In[28]:


df_hte['treatpost_hisp'] = df_hte.treatpost * df_hte.majority_hisp_pre
df_hte['treatpost_hisp_rural'] = df_hte.treatpost_hisp * df_hte.rural_pre
mod = PanelOLS.from_formula('avescores ~ + 1 + treatpost + treatpost_hisp + treatpost_hisp_rural + TimeEffects + EntityEffects', df_hte)
res = mod.fit(cov_type='clustered', cluster_entity=True)
print(res)


# In[27]:


df_hte[df_hte.majority_hisp_pre == 1][['district', 'type_description', 'majority_hisp_pre']]


# # Rural

# In[13]:


df_hte['treatpost_rural'] = df_hte.treatpost * df_hte.rural_pre
mod = PanelOLS.from_formula('avescores ~ + 1 + treatpost + treatpost_rural + rural_pre +  students_hisp + students_num  + TimeEffects + EntityEffects', df_hte)
res = mod.fit(cov_type='clustered', cluster_entity=True)
print(res)


# # Performance

# In[76]:


df_hte['treatpost_low'] = df_hte.treatpost * df_hte.low_avescores_pre
mod = PanelOLS.from_formula('avescores ~ + 1 + treatpost + treatpost_low + students_hisp + students_num  + TimeEffects + EntityEffects', df_hte)
res = mod.fit(cov_type='clustered', cluster_entity=True)
print(res)


# In[77]:


df_hte['treatpost_high'] = df_hte.treatpost * df_hte.high_avescores_pre
mod = PanelOLS.from_formula('avescores ~ + 1 + treatpost + treatpost_high+ students_hisp + students_num  + TimeEffects + EntityEffects', df_hte)
res = mod.fit(cov_type='clustered', cluster_entity=True)
print(res)


# # Teacher Turnover

# In[78]:


df_hte['treatpost_turnover'] = df_hte.treatpost * df_hte.high_turnover
mod = PanelOLS.from_formula('avescores ~ + 1 + treatpost + treatpost_turnover + students_hisp + students_num + TimeEffects + EntityEffects', df_hte)
res = mod.fit(cov_type='clustered', cluster_entity=True)
print(res)


# In[ ]:

