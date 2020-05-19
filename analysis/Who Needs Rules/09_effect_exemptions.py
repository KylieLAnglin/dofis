
# %%


import os
import sys

import matplotlib.mlab as mlab
import matplotlib.pyplot as plt
import numpy as np
import pandas as pd
import statsmodels.formula.api as smf
from linearmodels import PanelOLS
from openpyxl import load_workbook
from patsy import dmatrices

sys.path.append("../")
from library import start, analysis

# %%

data_path = start.data_path
table_path = start.table_path
data = pd.read_csv(os.path.join(data_path, 'clean', 'gdid_subject.csv'),
                   sep=",", low_memory=False)
data = data[data.doi]
print(data[(data.doi)].district.nunique())
print(data.doi_year.value_counts())
data.sample(5)

df = data.reset_index()
df['year'] = pd.to_datetime(df['year'], format='%Y')
# add column year to index
df = data.set_index(['year', 'campus'])
# swap indexes
df.index = df.index.swaplevel(0, 1)
df[['district', 'doi_year', 'treatpost']].sample(5)


# %% Generate interactions

def create_interactions(variable: str, data: pd.DataFrame):
    data['treatpost_' + variable] = data['treatpost'] * data[variable]
    data['yearpost_' + variable] = data['yearpost'] * data[variable]
    data['yearpre_' + variable] = data['yearpre'] * data[variable]

    for preyr in [5, 4, 3, 2]:
        data[variable + '_pre' + str(preyr)] = data[variable] * data['pre' + str(preyr)]
    for postyr in [1, 2, 3]:
        data[variable + '_post' + str(postyr)] = data[variable] * data['post' + str(postyr)]
    return data


def create_gdid_model_w_interactions(variable: str):
    gdid_model = 'score_std ~ + 1 + treatpost + '
    gdid_model = gdid_model + 'treatpost_' + variable + ' + '
    gdid_model = gdid_model + 'C(test_by_year) + '
    gdid_model = gdid_model + 'EntityEffects'
    print(gdid_model)
    return gdid_model


def create_linear_model_w_interactions(variable: str):
    linear_gdid_model = 'score_std ~ + 1 + treatpost + yearpost + yearpre + '
    linear_gdid_model = linear_gdid_model + 'treatpost_' + variable + ' + '
    linear_gdid_model = linear_gdid_model + 'yearpre_' + variable + ' + '
    linear_gdid_model = linear_gdid_model + 'yearpost_' + variable + ' + '
    linear_gdid_model = linear_gdid_model + ' + C(test_by_year) + '
    linear_gdid_model = linear_gdid_model + 'EntityEffects'

    return linear_gdid_model


def create_event_model_w_interactions(variable: str):
    event_study_model = 'score_std ~ + 1 + pre5 + pre4 + pre3 + pre2 + ' \
        'post1 + post2 + post3'
    for preyr in ["_pre5", "_pre4", "_pre3", "_pre2"]:
        event_study_model = event_study_model + ' + ' + variable + preyr
    for postyr in ["_post1", "_post2", "_post3"]:
        event_study_model = event_study_model + ' + ' + variable + postyr
    event_study_model = event_study_model + ' + C(test_by_year) + EntityEffects'
    print(event_study_model)

    return event_study_model



# %% Minutes Exemption


df = create_interactions('exempt_minutes', df)

# GDID
mod = PanelOLS.from_formula(create_gdid_model_w_interactions('exempt_minutes'),
                            df[df.math == 1])
res = mod.fit(cov_type='clustered', clusters=df[df.math == 1].district)
print(res)

# Linear GDID
mod = PanelOLS.from_formula(create_linear_model_w_interactions('exempt_minutes'),
                            df[df.math == 1])
res = mod.fit(cov_type='clustered', clusters=df[df.math == 1].district)
print(res)

# Event Study
mod = PanelOLS.from_formula(create_event_model_w_interactions('exempt_minutes'),
                            df[df.math == 1])
res = mod.fit(cov_type='clustered', clusters=df[df.math == 1].district)
print(res)


# %%

def results_col(data, col, var):

    df = create_interactions(var, data)

    linear_model = create_linear_model_w_interactions(var)

    mod = PanelOLS.from_formula(linear_model, df)
    res = mod.fit(cov_type='clustered', clusters=df.district)
    row = 3
    for coef in ['yearpre', 'treatpost[T.True]', 'yearpost',
                 'yearpre_' + var, 'treatpost_' + var, 'yearpost_' + var]:
        ws.cell(row=row, column=col).value = analysis.coef_with_stars(
            res.params[coef], res.pvalues[coef])
        row = row + 1
        ws.cell(row=row, column=col).value = analysis.format_se(
            res.std_errors[coef])
        row = row + 1
        wb.save(file)



# %%

file = start.table_path + 'effects_by_exemption_math.xlsx'
wb = load_workbook(file)
ws = wb.active

col = 2
for reg in ['exempt_firstday', 'exempt_minutes', 'exempt_lastday',
            'exempt_certification', 'exempt_classsize', 'exempt_probation',
            'exempt_servicedays', 'exempt_eval',
            'exempt_attendance', 'exempt_behavior']:
    results_col(data=df[df.math == 1], col=col, var=reg)
    col = col + 1

# %%
file = start.table_path + 'effects_by_exemption_reading.xlsx'
wb = load_workbook(file)
ws = wb.active

col = 2
for reg in ['exempt_firstday', 'exempt_minutes', 'exempt_lastday',
            'exempt_certification', 'exempt_classsize', 'exempt_probation',
            'exempt_servicedays', 'exempt_eval',
            'exempt_attendance', 'exempt_behavior']:
    results_col(data=df[df.reading == 1], col=col, var=reg)
    col = col + 1


# %%

exemption = 'exempt_firstday'

df = create_interactions(exemption, df)

# GDID
mod = PanelOLS.from_formula(create_gdid_model_w_interactions(exemption),
                            df[df.math == 1])
res = mod.fit(cov_type='clustered', clusters=df[df.math == 1].district)
print(res)


# Event Study
mod = PanelOLS.from_formula(create_event_model_w_interactions(exemption),
                            df[df.math == 1])
res = mod.fit(cov_type='clustered', clusters=df[df.math == 1].district)
print(res)


# %%
