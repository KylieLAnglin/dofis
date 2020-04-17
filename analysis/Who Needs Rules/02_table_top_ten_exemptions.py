#!/usr/bin/env python
# coding: utf-8

# In[4]:


import pandas as pd
import os
from openpyxl import load_workbook

data_path = '/Users/kylieleblancKylie/domino/dofis/data/clean'
table_path = '/Users/kylieleblancKylie/domino/dofis/results/Who Needs Rules/'


# In[5]:


data_path


# In[6]:


data = pd.read_csv((os.path.join(data_path, 'master_data_district.csv')))
data


# In[7]:


stubnames = sorted(
    set([match[0] for match in data.columns.str.findall(
    r'reg.*').values if match != [] ])
    )

proportion = []
number = []
for reg in stubnames:
    proportion.append(data[reg].mean())
    number.append(data[reg].sum())

regs = pd.DataFrame(
    {'law': stubnames,
     'proportion': proportion,
     'number': number
    }).sort_values(by = ['proportion'], ascending = False)
regs.head(10)


# In[9]:


from openpyxl import load_workbook

def df_to_excel(file, df, df_columns, start_col, start_row, all_ints = True):
    wb = load_workbook(file)
    ws = wb.active

    col_n = start_col
    for col in df_columns:
        row_n = start_row
        for ob in df[col]:
            ws.cell(row=row_n, column=col_n).value = ob
            row_n = row_n + 1
        col_n = col_n + 1
    wb.save(file)


# In[10]:


df_to_excel(file = os.path.join(table_path, 'table1_exemptions.xlsx'),
            df = regs, df_columns = ['law', 'proportion'],
            start_col = 4, start_row = 3)


# In[ ]:




