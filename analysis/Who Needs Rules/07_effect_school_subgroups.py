#!/usr/bin/env python
# coding: utf-8

# %%


import os
import sys

import matplotlib.mlab as mlab
import matplotlib.pyplot as plt
import numpy as np
import pandas as pd
import statsmodels.formula.api as smf
from linearmodels import PanelOLS
from openpyxl import load_workbook
from patsy import dmatrices

sys.path.append("../")
from library import analysis, start

pd.set_option('display.max_columns', None)


# %%


data_path = start.data_path
table_path = start.table_path
data = pd.read_csv(os.path.join(data_path, 'clean', 'gdid_subject.csv'),
                   sep=",", low_memory=False)
data = data[data.doi]
print(data[(data.doi)].district.nunique())
print(data.doi_year.value_counts())
data.sample(5)


# %%

data_pre = data[data.year == 2016]

fig, ax = plt.subplots(2, 2, sharey=False, sharex=False, figsize=(10, 10))

ax1 = ax[0, 0]
ax2 = ax[0, 1]
ax3 = ax[1, 0]
ax4 = ax[1, 1]


bins = 10

ax1.hist(data_pre.pre_hisp, bins, facecolor='gray', alpha=0.75)
ax1.set_xlabel('Percent Hispanic', fontsize=12)
for p in [.25, .5, .75, 1]:
    ax1.axvline(data_pre.pre_hisp.quantile(p), color='black', alpha=0.5)


ax2.hist(data_pre.pre_ell, bins, facecolor='gray', alpha=0.75)
ax2.set_xlabel('Percent ELL', fontsize=12)
for p in [.25, .5, .75, 1]:
    ax2.axvline(data_pre.pre_ell.quantile(p), color='black', alpha=0.5)

ax3.hist(data_pre.pre_num,  bins, facecolor='gray', alpha=0.75)
ax3.set_xlabel('Number of Students', fontsize=12)
for p in [.25, .5, .75, 1]:
    ax3.axvline(data_pre.pre_num.quantile(p), color='black', alpha=0.5)


ax4.hist(data_pre.pre_turnover.dropna(),  bins, facecolor='gray', alpha=0.75)
ax4.set_xlabel('Percent Turnover', fontsize=12)
for p in [.25, .5, .75, 1]:
    ax4.axvline(data_pre.pre_turnover.quantile(p), color='black', alpha=0.5)


fig.text(0.04, 0.5, 'Number of Schools', va='center',
         rotation='vertical', fontsize=12)
fig.suptitle('Pre-Implementation Characteristics', fontsize=14)
fig.subplots_adjust(top=0.95)
fig.savefig(table_path + 'Pre-Implementation Characteristics.png')

fig.text(0.04, 0.04, 'Notes: Characteristics are calculated for each school'
         ' in the year before implementation.'
         ' Turnover is caluclated at the district-level.')
fig.text(0.04, 0.02, 'The four horizontal lines represent the '
         '25th, 50th, 75th and 100th percentile.')


# %%

def print_quartile_stats(data: pd.DataFrame,
                         column_root: str, column_label: str):
    for quart in ['25', '50', '75', '100']:
        print(quart + "th Percentile Mean for " + column_label + ":")
        column_str = column_root + quart
        print(round(data.loc[data[column_str] == 1, column_root].mean(), 2))


print_quartile_stats(data_pre, 'pre_turnover', 'Teacher Turnover')
print_quartile_stats(data_pre, 'pre_avescore',
                     'Average Standardized Test Scores')
print_quartile_stats(data_pre, 'pre_hisp', 'Percent Hispanic Students')
print_quartile_stats(data_pre, 'pre_black', 'Percent Black Students')

# %%

# convert year to datetime
df = data.reset_index()
df['year'] = pd.to_datetime(df['year'], format='%Y')
# add column year to index
df = df.set_index(['year', 'campus'])
# swap indexes
df.index = df.index.swaplevel(0, 1)
df[['district', 'doi_year', 'treatpost']].sample(5, random_state=8)


# In[32]:


def results_table(data, col, var, q_vars):
    event_study_model = 'score_std ~ + 1 + pre5 + pre4 + pre3 + pre2 + ' \
        'post1 + post2 + post3 + '
    for preyr in [5, 4, 3, 2]:
        for p in q_vars:
            prevar = 'pre' + str(preyr) + var + str(p)
            oldprevar = 'pre' + str(preyr)
            oldcovar = var + str(p)
            data[prevar] = data[oldprevar] * data[oldcovar]
            event_study_model = event_study_model + prevar + ' + '
    postvars = []
    for postyr in [1, 2, 3]:
        postvars.append('post' + str(postyr))
        for p in q_vars:
            postvar = 'post' + str(postyr) + var + str(p)
            postvars.append(postvar)
            oldpostvar = 'post' + str(postyr)
            oldcovar = var + str(p)
            data[postvar] = data[oldpostvar] * data[oldcovar]
            event_study_model = event_study_model + postvar + ' + '
    event_study_model = event_study_model + ' C(test_by_year) + EntityEffects'
    print(event_study_model)
    # Event Study
    mod = PanelOLS.from_formula(event_study_model, data)
    res = mod.fit(cov_type='clustered', clusters=data.district)
    print(postvars)
    print(res)
    row = 3
    for coef in postvars:
        ws.cell(row=row, column=col).value = analysis.coef_with_stars(
            res.params[coef], res.pvalues[coef])
        row = row + 1
        ws.cell(row=row, column=col).value = analysis.format_se(
            res.std_errors[coef])
        row = row + 1

    wb.save(file)


# In[36]:
file = table_path + 'table5_hte_math.xlsx'
wb = load_workbook(file)
ws = wb.active

results_table(df[df.math == 1], 2, 'pre_', ['town', 'suburban', 'urban'])
results_table(df[df.math == 1], 3, 'pre_turnover', ['50', '75', '100'])
results_table(df[df.math == 1], 4, 'pre_avescore', ['50', '75', '100'])
results_table(df[df.math == 1], 5, 'pre_hisp', ['50', '75', '100'])
results_table(df[df.math == 1], 6, 'pre_black', ['50', '75', '100'])


# %% Reading

file = table_path + 'table6_hte_reading.xlsx'
wb = load_workbook(file)
ws = wb.active

results_table(df[df.reading == 1], 2, 'pre_', ['town', 'suburban', 'urban'])
results_table(df[df.reading == 1], 3, 'pre_turnover', ['50', '75', '100'])
results_table(df[df.reading == 1], 4, 'pre_avescore', ['50', '75', '100'])
results_table(df[df.reading == 1], 5, 'pre_hisp', ['50', '75', '100'])
results_table(df[df.reading == 1], 6, 'pre_black', ['50', '75', '100'])


# # Graphs

# In[60]:


def results_graph(data, var, q_vars, labels, saveas, title):
    event_study_model = 'score_std ~ + 1 + '
    for preyr in [5, 4, 3, 2]:
        for p in q_vars:
            prevar = 'pre' + str(preyr) + var + str(p)
            oldprevar = 'pre' + str(preyr)
            oldcovar = var + str(p)
            data[prevar] = data[oldprevar] * data[oldcovar]
            event_study_model = event_study_model + prevar + ' + '
    postvars = []
    for postyr in [1, 2, 3]:
        postvars.append('post' + str(postyr))
        for p in q_vars:
            postvar = 'post' + str(postyr) + var + str(p)
            postvars.append(postvar)
            oldpostvar = 'post' + str(postyr)
            oldcovar = var + str(p)
            data[postvar] = data[oldpostvar] * data[oldcovar]
            event_study_model = event_study_model + postvar + ' + '
    event_study_model = event_study_model + ' C(test_by_year) + EntityEffects'
    # Event Study
    mod = PanelOLS.from_formula(event_study_model, data)
    res = mod.fit(cov_type='clustered', clusters=data.district)

    fig, ax = plt.subplots(2, 2, sharey=False, sharex=False, figsize=(10, 10))

    ax1 = ax[0, 0]
    ax2 = ax[0, 1]
    ax3 = ax[1, 0]
    ax4 = ax[1, 1]

    for ax, p, q in zip([ax1, ax2, ax3, ax4], q_vars, labels):
        nonparametric = []
        nonparametric_se = []
        for coef in ['pre5' + var + p, 'pre4' + var + p, 'pre3' + var + p,
                     'pre2' + var + p, 'pre1' + var + p, 'post1' + var + p,
                     'post2' + var + p, 'post3' + var + p]:
            nonpar = 0
            nonpar_se = 0
            if coef != 'pre1' + var + p:
                nonpar = res.params[coef]
                nonpar_se = res.std_errors[coef]
            nonparametric.append(nonpar)
            nonparametric_se.append(nonpar_se)
        coef_df = pd.DataFrame({'coef': nonparametric,
                                'err': nonparametric_se,
                                'year': [-5, -4, -3, -2, -1, 1, 2, 3]
                                })
        coef_df['lb'] = coef_df.coef - (1.96*coef_df.err)
        coef_df['ub'] = coef_df.coef + (1.96*coef_df.err)
        coef_df['errsig'] = coef_df.err * 1.96

        # fig, ax = plt.subplots(figsize=(8, 5))

        coef_df.plot(x='year', y='coef', kind='bar',
                     ax=ax, color='none',
                     yerr='errsig', legend=False)
        ax.set_ylabel('')
        ax.set_xlabel('')
        ax.scatter(x=pd.np.arange(coef_df.shape[0]),
                   marker='s', s=120,
                   y=coef_df['coef'], color='black')
        ax.axhline(y=0, linestyle='--', color='black', linewidth=2)
        ax.xaxis.set_ticks_position('none')
        _ = ax.set_xticklabels(['Pre5', 'Pre4', 'Pre3', 'Pre2', 'Pre1',
                                'Post1', 'Post2', 'Post3'],
                               rotation=0)
        ax.set_ylim((-.5, .5))
        ax.set_title(q, fontsize=16)
    fig.savefig(table_path + saveas)
    fig.suptitle(title, fontsize='xx-large')


# %% Urbanicity

categories = ['rural', 'town', 'suburban', 'urban']
results_graph(df[df.math == 1], 'pre_', categories,
              ['Rural Schools', 'Town Schools', 'Suburban Schools',
               'Urban Schools'], 'Event Study Math Urbanicity.png',
              'Impact Estimates on Standardized Math Achievement '
              'by Urbanicity')


results_graph(df[df.reading == 1], 'pre_', categories,
              ['Rural Schools', 'Town Schools', 'Suburban Schools',
               'Urban Schools'],  'Event Study Reading Urbanicity.png',
              'Impact Estimates on Standardized Reading Achievement '
              'by Urbanicity')


# %% Turnover

results_graph(df[df.math == 1], 'pre_turnover', ['25', '50', '75', '100'],
              ['Q1 (Mean = 11%)', 'Q2 (Mean = 14%)',
               'Q3 (Mean = 18%)', 'Q4 (Mean = 25%)'],
              'Event Study Math Teacher Turnover.png',
              'Impact Estimates on Standardized Math Achievement '
              'by Teacher Turnover')


results_graph(df[df.reading == 1], 'pre_turnover', ['25', '50', '75', '100'],
              ['Q1 (Mean = 11%)', 'Q2 (Mean = 14%)',
               'Q3 (Mean = 18%)', 'Q4 (Mean = 25%)'],
              'Event Study Reading Teacher Turnover.png',
              'Impact Estimates on Standardized Reading Achievement '
              'by Teacher Turnover')


# %% Prior Achievement


results_graph(df[df.math == 1], 'pre_avescore', ['25', '50', '75', '100'],
              ['Q1 (-0.88 SD)', 'Q2 (-0.09 SD)',
               'Q3 (0.58 SD)', 'Q4 (1.62 SD)'],
              'Event Study Math Prior Achievement.png',
              'Impact Estimates on Standardized Math Achievement '
              'by Prior Achievement')


results_graph(df[df.reading == 1], 'pre_avescore', ['25', '50', '75', '100'],
              ['Q1 (-0.88 SD)', 'Q2 (-0.09 SD)',
               'Q3 (0.58 SD)', 'Q4 (1.62 SD)'],
              'Event Study Reading Prior Achievement.png',
              'Impact Estimates on Standardized Reading Achievement '
              'by Prior Achievement')

# %% Percent Hispanic
results_graph(df[df.math == 1], 'pre_hisp', ['25', '50', '75', '100'],
              ['Q1 (14%)', 'Q2 (31%)', 'Q3 (54%)', 'Q4 (86%)'],
              'Event Study Math Percent Hispanic.png',
              'Impact Estimates on Standardized Math Achievement '
              'by Perfect Hispanic Students')


results_graph(df[df.reading == 1], 'pre_hisp', ['25', '50', '75', '100'],
              ['Q1 (14%)', 'Q2 (31%)', 'Q3 (54%)', 'Q4 (86%)'],
              'Event Study Reading Percent Hispanic.png',
              'Impact Estimates on Standardized Reading Achievement '
              'by Perfect Hispanic Students')

# %% Percent Black


results_graph(df[df.math == 1], 'pre_black', ['25', '50', '75', '100'],
              ['Q1 (1%)', 'Q2 (4%)', 'Q3 (10%)', 'Q4 (31%)'],
              'Event Study Math Percent Black.png',
              'Impact Estimates on Standardized Math Achievement '
              'by Percent Black Students')

results_graph(df[df.reading == 1], 'pre_black', ['25', '50', '75', '100'],
              ['Q1 (1%)', 'Q2 (4%)', 'Q3 (10%)', 'Q4 (31%)'],
              'Event Study Reading Percent Black.png',
              'Impact Estimates on Standardized Reading Achievement '
              'by Percent Black Students')
