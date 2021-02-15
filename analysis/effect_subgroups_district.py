#!/usr/bin/env python
# coding: utf-8

# %%


import os
import sys

import matplotlib.mlab as mlab
import matplotlib.pyplot as plt
import numpy as np
import pandas as pd
import statsmodels.formula.api as smf
from linearmodels import PanelOLS
from openpyxl import load_workbook
from patsy import dmatrices

sys.path.append("../")
from library import analysis, start

pd.set_option("display.max_columns", None)


# %%


data = pd.read_csv(
    os.path.join(start.data_path, "clean", "master_data_district.csv"),
    sep=",",
    low_memory=False,
)
data = data[data.doi]
data = data[data.doi_year > 2016]
# data = data[data.analytic_sample]
print(data[data.year == 2016].doi_year.value_counts())

# data = pd.get_dummies(data=data, prefix="yr", columns=["year"])

data.sample()

# %%

# convert year to datetime
df = data.reset_index()
df["year_index"] = pd.to_datetime(df["year"], format="%Y")
df["district_index"] = df.district
# add column year to index
df = df.set_index(["district_index", "year_index"])
df[
    [
        "year",
        "doi_year",
        "treatpost",
        "yearpost",
        "post1",
        "math",
        "reading",
        "elem_math",
    ]
].sample(5)


def results_table(data, col, outcome, var, q_vars):
    event_study_model = outcome + (
        " ~ + 1 + pre5 + pre4 + pre3 + pre2 + " "post1 + post2 + post3 + "
    )
    for preyr in [5, 4, 3, 2]:
        for p in q_vars:
            prevar = "pre" + str(preyr) + var + str(p)
            oldprevar = "pre" + str(preyr)
            oldcovar = var + str(p)
            data[prevar] = data[oldprevar] * data[oldcovar]
            event_study_model = event_study_model + prevar + " + "
    postvars = []
    for postyr in [1, 2, 3]:
        postvars.append("post" + str(postyr))
        for p in q_vars:
            postvar = "post" + str(postyr) + var + str(p)
            postvars.append(postvar)
            oldpostvar = "post" + str(postyr)
            oldcovar = var + str(p)
            data[postvar] = data[oldpostvar] * data[oldcovar]
            event_study_model = event_study_model + postvar + " + "
    event_study_model = event_study_model + " C(year) + EntityEffects"
    print(event_study_model)
    # Event Study
    mod = PanelOLS.from_formula(event_study_model, data)
    res = mod.fit(cov_type="clustered", clusters=data.district)
    print(postvars)
    print(res)
    row = 3
    for coef in postvars:
        ws.cell(row=row, column=col).value = analysis.coef_with_stars(
            res.params[coef], res.pvalues[coef]
        )
        row = row + 1
        ws.cell(row=row, column=col).value = analysis.format_se(res.std_errors[coef])
        row = row + 1


# In[36]:
file = table_path + "hte_math_district.xlsx"
wb = load_workbook(file)
ws = wb.active

results_table(
    data=df, col=2, outcome="math", var="pre_", q_vars=["town", "suburban", "urban"]
)

# results_table(data = df[df.math == 1], col=2, "pre_", ["town", "suburban", "urban"])
# results_table(df[df.math == 1], 3, "pre_turnover", ["50", "75", "100"])
# results_table(df[df.math == 1], 4, "pre_avescore", ["50", "75", "100"])
# results_table(df[df.math == 1], 5, "pre_hisp", ["50", "75", "100"])
# results_table(df[df.math == 1], 6, "pre_black", ["50", "75", "100"])


# %% Reading

file = table_path + "table6_hte_reading.xlsx"
wb = load_workbook(file)
ws = wb.active

results_table(df[df.reading == 1], 2, "pre_", ["town", "suburban", "urban"])
results_table(df[df.reading == 1], 3, "pre_turnover", ["50", "75", "100"])
results_table(df[df.reading == 1], 4, "pre_avescore", ["50", "75", "100"])
results_table(df[df.reading == 1], 5, "pre_hisp", ["50", "75", "100"])
results_table(df[df.reading == 1], 6, "pre_black", ["50", "75", "100"])


# # Graphs

# In[60]:


def results_graph(data, var, q_vars, labels, saveas, title):
    event_study_model = "score_std ~ + 1 + "
    for preyr in [5, 4, 3, 2]:
        for p in q_vars:
            prevar = "pre" + str(preyr) + var + str(p)
            oldprevar = "pre" + str(preyr)
            oldcovar = var + str(p)
            data[prevar] = data[oldprevar] * data[oldcovar]
            event_study_model = event_study_model + prevar + " + "
    postvars = []
    for postyr in [1, 2, 3]:
        postvars.append("post" + str(postyr))
        for p in q_vars:
            postvar = "post" + str(postyr) + var + str(p)
            postvars.append(postvar)
            oldpostvar = "post" + str(postyr)
            oldcovar = var + str(p)
            data[postvar] = data[oldpostvar] * data[oldcovar]
            event_study_model = event_study_model + postvar + " + "
    event_study_model = event_study_model + " C(test_by_year) + EntityEffects"
    # Event Study
    mod = PanelOLS.from_formula(event_study_model, data)
    res = mod.fit(cov_type="clustered", clusters=data.district)

    fig, ax = plt.subplots(2, 2, sharey=False, sharex=False, figsize=(10, 10))

    ax1 = ax[0, 0]
    ax2 = ax[0, 1]
    ax3 = ax[1, 0]
    ax4 = ax[1, 1]

    for ax, p, q in zip([ax1, ax2, ax3, ax4], q_vars, labels):
        nonparametric = []
        nonparametric_se = []
        for coef in [
            "pre5" + var + p,
            "pre4" + var + p,
            "pre3" + var + p,
            "pre2" + var + p,
            "pre1" + var + p,
            "post1" + var + p,
            "post2" + var + p,
            "post3" + var + p,
        ]:
            nonpar = 0
            nonpar_se = 0
            if coef != "pre1" + var + p:
                nonpar = res.params[coef]
                nonpar_se = res.std_errors[coef]
            nonparametric.append(nonpar)
            nonparametric_se.append(nonpar_se)
        coef_df = pd.DataFrame(
            {
                "coef": nonparametric,
                "err": nonparametric_se,
                "year": [-5, -4, -3, -2, -1, 1, 2, 3],
            }
        )
        coef_df["lb"] = coef_df.coef - (1.96 * coef_df.err)
        coef_df["ub"] = coef_df.coef + (1.96 * coef_df.err)
        coef_df["errsig"] = coef_df.err * 1.96

        # fig, ax = plt.subplots(figsize=(8, 5))

        coef_df.plot(
            x="year",
            y="coef",
            kind="bar",
            ax=ax,
            color="none",
            yerr="errsig",
            legend=False,
        )
        ax.set_ylabel("")
        ax.set_xlabel("")
        ax.scatter(
            x=pd.np.arange(coef_df.shape[0]),
            marker="s",
            s=120,
            y=coef_df["coef"],
            color="black",
        )
        ax.axhline(y=0, linestyle="--", color="black", linewidth=2)
        ax.xaxis.set_ticks_position("none")
        _ = ax.set_xticklabels(
            ["Pre5", "Pre4", "Pre3", "Pre2", "Pre1", "Post1", "Post2", "Post3"],
            rotation=0,
        )
        ax.set_ylim((-0.5, 0.5))
        ax.set_title(q, fontsize=16)
    fig.suptitle(title, fontsize="xx-large")
    fig.savefig(table_path + saveas)


# %% Urbanicity

categories = ["rural", "town", "suburban", "urban"]
results_graph(
    df[df.math == 1],
    "pre_",
    categories,
    ["Rural Schools", "Town Schools", "Suburban Schools", "Urban Schools"],
    "Event Study Math Urbanicity.png",
    "Standardized Math Achievement " "by Urbanicity",
)


results_graph(
    df[df.reading == 1],
    "pre_",
    categories,
    ["Rural Schools", "Town Schools", "Suburban Schools", "Urban Schools"],
    "Event Study Reading Urbanicity.png",
    "Standardized Reading Achievement " "by Urbanicity",
)


# %% Turnover

results_graph(
    df[df.math == 1],
    "pre_turnover",
    ["25", "50", "75", "100"],
    ["Q1 (Mean = 11%)", "Q2 (Mean = 14%)", "Q3 (Mean = 18%)", "Q4 (Mean = 25%)"],
    "Event Study Math Teacher Turnover.png",
    "Impact Estimates on Standardized Math Achievement " "by Teacher Turnover",
)


results_graph(
    df[df.reading == 1],
    "pre_turnover",
    ["25", "50", "75", "100"],
    ["Q1 (Mean = 11%)", "Q2 (Mean = 14%)", "Q3 (Mean = 18%)", "Q4 (Mean = 25%)"],
    "Event Study Reading Teacher Turnover.png",
    "Impact Estimates on Standardized Reading Achievement " "by Teacher Turnover",
)


# %% Prior Achievement

for q in ["25", "50", "75", "100"]:
    print(df[df["pre_avescore" + q] == 1].avescores.mean())


results_graph(
    df[df.math == 1],
    "pre_avescore",
    ["25", "50", "75", "100"],
    ["Q1 (-0.78 SD)", "Q2 (-0.20 SD)", "Q3 (0.25 SD)", "Q4 (1.18 SD)"],
    "Event Study Math Prior Achievement.png",
    "Standardized Math Achievement",
)


results_graph(
    df[df.reading == 1],
    "pre_avescore",
    ["25", "50", "75", "100"],
    ["Q1 (-0.78 SD)", "Q2 (-0.20 SD)", "Q3 (0.25 SD)", "Q4 (1.18 SD)"],
    "Event Study Reading Prior Achievement.png",
    "Standardized Reading Achievement",
)

# %% Percent Hispanic
for q in ["25", "50", "75", "100"]:
    print(df[df["pre_hisp" + q] == 1].students_hisp.mean())


results_graph(
    df[df.math == 1],
    "pre_hisp",
    ["25", "50", "75", "100"],
    ["Q1 (15%)", "Q2 (35%)", "Q3 (60%)", "Q4 (91%)"],
    "Event Study Math Percent Hispanic.png",
    "Standardized Math Achievement",
)


results_graph(
    df[df.reading == 1],
    "pre_hisp",
    ["25", "50", "75", "100"],
    ["Q1 (15%)", "Q2 (35%)", "Q3 (60%)", "Q4 (91%)"],
    "Event Study Reading Percent Hispanic.png",
    "Standardized Reading Achievement",
)

# %% Percent Black
for q in ["25", "50", "75", "100"]:
    print(df[df["pre_black" + q] == 1].students_black.mean())

results_graph(
    df[df.math == 1],
    "pre_black",
    ["25", "50", "75", "100"],
    ["Q1 (1%)", "Q2 (3%)", "Q3 (10%)", "Q4 (32%)"],
    "Event Study Math Percent Black.png",
    "Standardized Math Achievement",
)

results_graph(
    df[df.reading == 1],
    "pre_black",
    ["25", "50", "75", "100"],
    ["Q1 (1%)", "Q2 (3%)", "Q3 (10%)", "Q4 (32%)"],
    "Event Study Reading Percent Black.png",
    "Standardized Reading Achievement",
)


# %%
