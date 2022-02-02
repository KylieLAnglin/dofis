# %%

import pandas as pd
import numpy as np
import pandas as pd
from openpyxl import load_workbook
import scipy

from dofis import start
from dofis.analysis.library import analysis

# %%
results_path = start.TABLE_PATH + "results_subjects_raw.xlsx"
results = pd.read_excel(results_path).set_index("outcome")


file_path = start.TABLE_PATH + "Aggregated Impact of DOI Status by Subject.xlsx"
wb = load_workbook(file_path)
ws = wb.active

n = 8
col = 2
row = 4
for outcome in [
    "m_3rd_yr15std",
    "m_4th_yr15std",
    "m_5th_yr15std",
    "m_6th_yr15std",
    "m_7th_yr15std",
    "m_8th_yr15std",
    "alg_yr15std",
    "biology_yr15std",
]:
    te = results.loc[outcome, "te"]
    se = results.loc[outcome, "se"]
    pvalue = (
        scipy.stats.t.sf(np.abs(te / se), n - 1) * 2
    )  # two-sided pvalue = Prob(abs(t)>tt)

    ws.cell(row=row, column=col).value = analysis.coef_with_stars(
        te, pvalue=pvalue, n_tests=7, digits=2
    )
    row = row + 1
    ws.cell(row=row, column=col).value = analysis.format_se(se, 2)
    row = row + 1

col = 4
row = 5
for outcome in [
    "r_3rd_yr15std",
    "r_4th_yr15std",
    "r_5th_yr15std",
    "r_6th_yr15std",
    "r_7th_yr15std",
    "r_8th_yr15std",
    "eng1_yr15std",
    "us_yr15std",
]:
    te = results.loc[outcome, "te"]
    se = results.loc[outcome, "se"]
    pvalue = (
        scipy.stats.t.sf(np.abs(te / se), n - 1) * 2
    )  # two-sided pvalue = Prob(abs(t)>tt)

    ws.cell(row=row, column=col).value = analysis.coef_with_stars(
        te, pvalue=pvalue, n_tests=7, digits=2
    )
    row = row + 1
    ws.cell(row=row, column=col).value = analysis.format_se(se, 2)
    row = row + 1

wb.save(file_path)
