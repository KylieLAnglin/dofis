#!/usr/bin/env python
# coding: utf-8

# %%


import os
import sys

import numpy as np
import pandas as pd
import statsmodels.formula.api as smf
import scipy
from openpyxl import load_workbook


from dofis.analysis.library import analysis, characteristics, regulations, tables
from dofis import start

# %%


DATA_PATH = start.DATA_PATH
TABLE_PATH = start.TABLE_PATH
data = pd.read_csv(
    os.path.join(start.DATA_PATH, "clean", "master_data_district.csv"), sep=","
)
data = data[data.year == 2016]
data["teachers_turnover_ratio_d"] = data.teachers_turnover_ratio_d / 100
data.group.value_counts()

print("Number ineligble districts:", len(data[data.group == "ineligible"]))
data[data.group != "ineligible"]
print(
    "Number DOIs with missing implementation year:",
    len(data[(data.doi == 1) & (data.doi_year.isnull())]),
)

# %%
# %%
FILE = start.TABLE_PATH + "Characteristics of Implementation Groups.xlsx"
wb = load_workbook(FILE)
ws = wb.active


# %% Sample sizes
ws.cell(row=3, column=2).value = "N = " + str(len(data[data.group == "opt-out"]))
ws.cell(row=3, column=3).value = "N = " + str(len(data[data.group == "2017"]))
ws.cell(row=3, column=4).value = "N = " + str(len(data[data.group == "2018"]))
ws.cell(row=3, column=5).value = "N = " + str(len(data[data.group == "2019"]))
ws.cell(row=3, column=6).value = "N = " + str(len(data[data.group == "2020+"]))

# %%


def mean_and_sd_to_table(
    df: pd.DataFrame, variable: str, start_row: int, col: int, round_value: int = 2
):
    ws.cell(row=start_row, column=col).value = str(
        round(df[variable].mean(), round_value)
    )
    # ws.cell(row=start_row + 1, column=col).value = (
    #     "[" + str(round(df[variable].std(), round_value)) + "]"
    # )
    ws.cell(row=start_row + 1, column=col).value = (
        "("
        + str(round(scipy.stats.sem(df[variable], nan_policy="omit"), round_value))
        + ")"
    )


for group, col in zip(
    [
        "2017",
        "2018",
        "2019",
        "2020+",
        "opt-out",
    ],
    [2, 3, 4, 5, 6],
):
    row = 4
    mean_and_sd_to_table(
        df=data[data.group == group], variable="type_urban", start_row=row, col=col
    )
    row = row + 2

    mean_and_sd_to_table(
        df=data[data.group == group], variable="type_suburban", start_row=row, col=col
    )
    row = row + 2

    mean_and_sd_to_table(
        df=data[data.group == group], variable="type_town", start_row=row, col=col
    )
    row = row + 2

    mean_and_sd_to_table(
        df=data[data.group == group], variable="type_rural", start_row=row, col=col
    )
    row = row + 2

    mean_and_sd_to_table(
        df=data[data.group == group],
        variable="teachers_exp_ave",
        start_row=row,
        col=col,
        round_value=1,
    )
    row = row + 2

    mean_and_sd_to_table(
        df=data[data.group == group],
        variable="teachers_tenure_ave",
        start_row=row,
        col=col,
        round_value=1,
    )
    row = row + 2

    mean_and_sd_to_table(
        df=data[data.group == group],
        variable="teachers_turnover_ratio_d",
        start_row=row,
        col=col,
        round_value=2,
    )
    row = row + 2

    mean_and_sd_to_table(
        df=data[data.group == group],
        variable="class_size_5",
        start_row=row,
        col=col,
        round_value=1,
    )
    row = row + 2

    mean_and_sd_to_table(
        df=data[data.group == group],
        variable="students_hisp",
        start_row=row,
        col=col,
        round_value=2,
    )
    row = row + 2

    mean_and_sd_to_table(
        df=data[data.group == group],
        variable="students_black",
        start_row=row,
        col=col,
        round_value=2,
    )
    row = row + 2

    mean_and_sd_to_table(
        df=data[data.group == group],
        variable="students_white",
        start_row=row,
        col=col,
        round_value=2,
    )
    row = row + 2

    mean_and_sd_to_table(
        df=data[data.group == group],
        variable="students_frpl",
        start_row=row,
        col=col,
        round_value=2,
    )
    row = row + 2

    mean_and_sd_to_table(
        df=data[data.group == group],
        variable="avescores",
        start_row=row,
        col=col,
        round_value=2,
    )
    row = row + 2


wb.save(FILE)

# %%
