# %%

import os
import sys

import numpy as np
import pandas as pd
import statsmodels.formula.api as smf
import scipy
from openpyxl import load_workbook


from dofis.analysis.library import analysis, characteristics, regulations, tables
from dofis import start

# %%
DATA_PATH = start.DATA_PATH
TABLE_PATH = start.TABLE_PATH
data = pd.read_csv(
    os.path.join(start.DATA_PATH, "clean", "master_data_district.csv"), sep=","
)
data = data[data.year == 2021]
data = data[data.group != "charter"]

list_of_demographics = [
    "temp",
    "students_num",
    "students_hisp_num",
    "students_white_num",
    "students_black_num",
    "students_asian_num",
    "students_frpl_num",
    "students_sped_num",
    "students_rural_num",
    "students_town_num",
    "students_suburban_num",
    "students_urban_num",
]

# %%

total_number_students = data.students_num.sum()
# %% Calculate number students by geography
data["students_rural_num"] = data.students_num * data.type_rural
data["students_town_num"] = data.students_num * data.type_town
data["students_suburban_num"] = data.students_num * data.type_suburban
data["students_urban_num"] = data.students_num * data.type_urban

# %%

data["temp"] = 1
totals_wide = data[list_of_demographics].groupby("temp").sum()
totals = totals_wide.T
totals = totals.reset_index().rename(
    columns={"index": "demographic", 1: "students_in_public_schools"}
)
# %%
doi_totals = data[list_of_demographics][data.doi == 1].groupby("temp").sum()
doi_totals = doi_totals.T
doi_totals = doi_totals.reset_index().rename(
    columns={"index": "demographic", 1: "students_in_dois"}
)
table_dem = totals.merge(doi_totals, left_on="demographic", right_on="demographic")
table_dem["prop_in_doi"] = (
    table_dem.students_in_dois / table_dem.students_in_public_schools
)
table_dem = table_dem.set_index("demographic")


# %%
# %%


FILE = start.TABLE_PATH + "Proportion Students in DOIs.xlsx"
wb = load_workbook(FILE)
ws = wb.active

# %% Sample sizes
groups_for_table = [
    "students_asian_num",
    "students_black_num",
    "students_hisp_num",
    "students_white_num",
    "students_frpl_num",
    "students_sped_num",
    "students_rural_num",
    "students_town_num",
    "students_suburban_num",
    "students_urban_num",
]

row = 2
col = 2
for group in groups_for_table:
    ws.cell(row=row, column=2).value = round(table_dem.loc[group, "prop_in_doi"], 2)
    row = row + 1

# %%
cert_totals = data[list_of_demographics][data.reg21_003 == 1].groupby("temp").sum()
cert_totals = cert_totals.T
cert_totals = cert_totals.reset_index().rename(
    columns={"index": "demographic", 1: "students_in_dois"}
)
table_cert = totals.merge(cert_totals, left_on="demographic", right_on="demographic")
table_cert["prop_in_doi"] = (
    table_cert.students_in_dois / table_cert.students_in_public_schools
)
table_cert = table_cert.set_index("demographic")
# %%
row = 2
for group in groups_for_table:
    ws.cell(row=row, column=3).value = round(table_cert.loc[group, "prop_in_doi"], 2)
    row = row + 1

wb.save(FILE)

# %%
class_totals = data[list_of_demographics][data.reg25_112 == 1].groupby("temp").sum()
class_totals = class_totals.T
class_totals = class_totals.reset_index().rename(
    columns={"index": "demographic", 1: "students_in_dois"}
)
class_totals = totals.merge(class_totals, left_on="demographic", right_on="demographic")
class_totals["prop_in_doi"] = (
    class_totals.students_in_dois / class_totals.students_in_public_schools
)
class_totals = class_totals.set_index("demographic")
# %%
row = 2
for group in groups_for_table:
    ws.cell(row=row, column=4).value = round(class_totals.loc[group, "prop_in_doi"], 2)
    row = row + 1

wb.save(FILE)

# %%
time_totals = data[list_of_demographics][data.reg25_081 == 1].groupby("temp").sum()
time_totals = time_totals.T
time_totals = time_totals.reset_index().rename(
    columns={"index": "demographic", 1: "students_in_dois"}
)
time_totals = totals.merge(time_totals, left_on="demographic", right_on="demographic")
time_totals["prop_in_doi"] = (
    time_totals.students_in_dois / time_totals.students_in_public_schools
)
time_totals = time_totals.set_index("demographic")
# %%
row = 2
for group in groups_for_table:
    ws.cell(row=row, column=5).value = round(time_totals.loc[group, "prop_in_doi"], 2)
    row = row + 1

wb.save(FILE)
# %%
attend_totals = data[list_of_demographics][data.reg25_092 == 1].groupby("temp").sum()
attend_totals = attend_totals.T
attend_totals = attend_totals.reset_index().rename(
    columns={"index": "demographic", 1: "students_in_dois"}
)
attend_totals = totals.merge(
    attend_totals, left_on="demographic", right_on="demographic"
)
attend_totals["prop_in_doi"] = (
    attend_totals.students_in_dois / attend_totals.students_in_public_schools
)
attend_totals = attend_totals.set_index("demographic")
# %%
row = 2
for group in groups_for_table:
    ws.cell(row=row, column=6).value = round(attend_totals.loc[group, "prop_in_doi"], 2)
    row = row + 1

wb.save(FILE)

# %%
