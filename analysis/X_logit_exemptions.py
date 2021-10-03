#!/usr/bin/env python
# coding: utf-8

# In[96]:


import os
import sys

import pandas as pd
import statsmodels.formula.api as sm
from openpyxl import load_workbook

from dofis import start
from dofis.analysis.library import analysis, characteristics, regulations, tables

pd.set_option("display.max_columns", None)


# In[97]:
data = pd.read_csv(
    os.path.join(start.DATA_PATH, "clean", "master_data_district.csv"), sep=","
)
data = data[(data.doi == 1)]
data = data[(data.year == 2019)]
data.sample(5)


# In[101]:


def model_to_excel(data, y, x_list, file, start_row, start_col):
    wb = load_workbook(file)
    ws = wb.active

    col_n = start_col
    row_n = start_row

    # create formula
    i = 1
    x_str = ""
    for x in x_list:
        if i == 1:
            x_str = x_str + x
        if i > 1:
            x_str = x_str + " + " + x
        i += 1
    x_str = x_str + "- 1"
    formula = y + " ~ " + x_str

    # run
    result = smf.logit(formula=formula, data=data).fit()
    # to excel
    for x in x_list:
        p = result.pvalues[x]
        coef = result.params[x].round(2)
        if p >= 0.05:
            coef = str(coef)
        if p < 0.05 and p > 0.01:
            coef = str(coef) + "*"
        if p < 0.01 and p > 0.001:
            coef = str(coef) + "**"
        if p < 0.001:
            coef = str(coef) + "***"
        ws.cell(row=row_n, column=col_n).value = coef
        row_n = row_n + 1
        ws.cell(row=row_n, column=col_n).value = "(" + str(result.bse[x].round(2)) + ")"
        row_n = row_n + 1

    ws.cell(row=row_n, column=col_n).value = result.nobs
    row_n = row_n + 1
    ws.cell(row=row_n, column=col_n).value = result.llf.round(2)
    row_n = row_n + 1
    ws.cell(row=row_n, column=col_n).value = result.prsquared.round(2)
    wb.save(file)

    print(result)

    return result.summary()


# In[102]:


x_list = [
    "pre_rural",
    "pre_town",
    "pre_suburban",
    "pre_turnover",
    "pre_tenure",
    "pre_ratio",
    "pre_hisp",
    "pre_black",
    "pre_frpl",
    "pre_avescore",
]


# %% Schedules


tables.model_to_excel(
    data,
    y="reg25_0811",
    x_list=x_list,
    file=start.TABLE_PATH + "logit_exemptions.xlsx",
    start_row=4,
    start_col=2,
)
tables.model_to_excel(
    data,
    y="reg25_081",
    x_list=x_list,
    file=start.TABLE_PATH + "logit_exemptions.xlsx",
    start_row=4,
    start_col=3,
)
tables.model_to_excel(
    data,
    y="reg25_0812",
    x_list=x_list,
    file=start.TABLE_PATH + "logit_exemptions.xlsx",
    start_row=4,
    start_col=4,
)


# %% Certification and class sizes


tables.model_to_excel(
    data,
    y="reg21_003",
    x_list=x_list,
    file=start.TABLE_PATH + "logit_exemptions.xlsx",
    start_row=4,
    start_col=5,
)
tables.model_to_excel(
    data,
    y="reg25_112",
    x_list=x_list,
    file=start.TABLE_PATH + "logit_exemptions.xlsx",
    start_row=4,
    start_col=6,
)
tables.model_to_excel(
    data,
    y="reg25_111",
    x_list=x_list,
    file=start.TABLE_PATH + "logit_exemptions.xlsx",
    start_row=4,
    start_col=7,
)

# %% Contracts


tables.model_to_excel(
    data,
    y="reg21_102",
    x_list=x_list,
    file=start.TABLE_PATH + "logit_exemptions.xlsx",
    start_row=4,
    start_col=8,
)
tables.model_to_excel(
    data,
    y="reg21_401",
    x_list=x_list,
    file=start.TABLE_PATH + "logit_exemptions.xlsx",
    start_row=4,
    start_col=9,
)
tables.model_to_excel(
    data,
    y="reg21_352",
    x_list=x_list,
    file=start.TABLE_PATH + "logit_exemptions.xlsx",
    start_row=4,
    start_col=10,
)
tables.model_to_excel(
    data,
    y="reg21_354",
    x_list=x_list,
    file=start.TABLE_PATH + "logit_exemptions.xlsx",
    start_row=4,
    start_col=11,
)

# %% Behavior


tables.model_to_excel(
    data,
    y="reg25_092",
    x_list=x_list,
    file=start.TABLE_PATH + "logit_exemptions.xlsx",
    start_row=4,
    start_col=12,
)
tables.model_to_excel(
    data,
    y="reg37_0012",
    x_list=x_list,
    file=start.TABLE_PATH + "logit_exemptions.xlsx",
    start_row=4,
    start_col=13,
)
tables.model_to_excel(
    data,
    y="reg25_036",
    x_list=x_list,
    file=start.TABLE_PATH + "logit_exemptions.xlsx",
    start_row=4,
    start_col=14,
)
