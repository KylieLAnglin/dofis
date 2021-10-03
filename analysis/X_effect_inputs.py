# %%
import pandas as pd
import statsmodels.formula.api as smf
from openpyxl import load_workbook


from dofis import start
from dofis.analysis.library import characteristics
from dofis.analysis.library import analysis


INPUTS = [
    "class_size_k",
    "class_size_1",
    "class_size_2",
    "class_size_3",
    "class_size_4",
    "class_size_5",
    "perf_attendance",
    "perf_stuattend",
    "perf_studays",
    "stu_teach_ratio",
    "days_max",
    "days_mean",
    "days_min",
    "teachers_certified",
    "teachers_uncertified",
    "teachers_secondary_math_certified",
    "teachers_secondary_math_uncertified",
    "teachers_secondary_math_outoffield",
    "teachers_secondary_science_outoffield",
    "teachers_secondary_science_uncertified",
    "teachers_secondary_cte_uncertified",
    "teachers_secondary_cte_outoffield",
]


COLUMNS = (
    ["year", "doi_year", "campus", "campname", "district", "distname"]
    + characteristics.POTENTIAL_COVARIATES
    + INPUTS
    + [
        "exempt_minutes",
        "exempt_certification",
        "exempt_classsize",
        "exempt_servicedays",
    ]
)

# %%
data_school = pd.read_csv(
    os.path.join(start.DATA_PATH, "clean", "master_data_school.csv"),
    sep=",",
    low_memory=False,
)
data_school = data_school[data_school.doi == 1]
data_school = data_school[COLUMNS]
data_school["const"] = 1

data_school = data_school.dropna(subset=characteristics.POTENTIAL_COVARIATES)
data_school = data_school.set_index("campus")

# %%
pre_data = data_school[data_school.year == 2016][characteristics.POTENTIAL_COVARIATES]
pre_data = pre_data.add_prefix("pre_")

data_school = data_school.merge(pre_data, how="left", left_index=True, right_index=True)
data_school = data_school.dropna(subset=list(data_school.filter(like="pre_").columns))


# %%


# %%

matching_variables = analysis.double_covariate_selection(
    "exempt_certification",
    "teachers_uncertified",
    data_school,
    list(data_school.filter(like="pre_").columns),
    alpha=0.005,
)
matching_variables


# %%


def effect_input(
    outcome: str,
    treatment: str,
    covariates: list,
    data: pd.DataFrame,
    table: str,
    col: int,
):
    model = outcome + " ~ 1 + " + treatment + " + "
    for var in covariates:
        model = model + " + " + var

    data = data.dropna(subset=[outcome] + [treatment] + covariates)

    mod = smf.ols(formula=model, data=data)
    res = mod.fit(cov_type="cluster", cov_kwds={"groups": data["district"]})
    print(res.summary())

    file = start.TABLE_PATH + table
    wb = load_workbook(file)
    ws = wb.active

    # Estimates
    ws.cell(row=4, column=col).value = analysis.coef_with_stars(
        res.params[treatment], res.pvalues[treatment], digits=3
    )
    ws.cell(row=5, column=col).value = analysis.format_se(res.bse[treatment], digits=3)

    # Sample sizes
    ws.cell(row=9, column=col).value = len(data[data[treatment] == 1])
    ws.cell(row=10, column=col).value = data[data[treatment] == 1].district.nunique()
    ws.cell(row=11, column=col).value = len(data[data[treatment] == 0])
    ws.cell(row=12, column=col).value = data[data[treatment] == 0].district.nunique()

    wb.save(file)


for year, col in zip([2017, 2018, 2019], [2, 3, 4]):
    effect_input(
        outcome="teachers_uncertified",
        treatment="exempt_certification",
        covariates=list(data_school.filter(like="pre_").columns),
        data=data_school[(data_school.year == year) & (data_school.doi_year <= year)],
        table="effect_uncertified_teachers.xlsx",
        col=col,
    )

for year, col in zip([2017, 2018, 2019], [5, 6, 7]):
    effect_input(
        outcome="teachers_secondary_cte_uncertified",
        treatment="exempt_certification",
        covariates=list(data_school.filter(like="pre_").columns),
        data=data_school[(data_school.year == year) & (data_school.doi_year <= year)],
        table="effect_uncertified_teachers.xlsx",
        col=col,
    )

for year, col in zip([2017, 2018, 2019], [8, 9, 10]):
    effect_input(
        outcome="teachers_secondary_math_uncertified",
        treatment="exempt_certification",
        covariates=list(data_school.filter(like="pre_").columns),
        data=data_school[(data_school.year == year) & (data_school.doi_year <= year)],
        table="effect_uncertified_teachers.xlsx",
        col=col,
    )

for year, col in zip([2017, 2018, 2019], [11, 12, 13]):
    effect_input(
        outcome="teachers_secondary_science_uncertified",
        treatment="exempt_certification",
        covariates=list(data_school.filter(like="pre_").columns),
        data=data_school[(data_school.year == year) & (data_school.doi_year <= year)],
        table="effect_uncertified_teachers.xlsx",
        col=col,
    )

for year, col in zip([2017, 2018, 2019], [2, 3, 4]):
    effect_input(
        outcome="teachers_secondary_science_outoffield",
        treatment="exempt_certification",
        covariates=list(data_school.filter(like="pre_").columns),
        data=data_school[(data_school.year == year) & (data_school.doi_year <= year)],
        table="effect_outoffield_teachers.xlsx",
        col=col,
    )

for year, col in zip([2017, 2018, 2019], [5, 6, 7]):
    effect_input(
        outcome="teachers_secondary_math_outoffield",
        treatment="exempt_certification",
        covariates=list(data_school.filter(like="pre_").columns),
        data=data_school[(data_school.year == year) & (data_school.doi_year <= year)],
        table="effect_outoffield_teachers.xlsx",
        col=col,
    )
