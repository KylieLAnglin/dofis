import pandas as pd
import numpy as np
import pandas as pd
from openpyxl import load_workbook
import scipy

from dofis import start
from dofis.analysis.library import analysis

# %%
results_path = start.TABLE_PATH + "/results_inputs_raw.xlsx"
results = pd.read_excel(results_path).set_index("outcome")


file_path = start.TABLE_PATH + "Aggregated Impact of DOI Status on Inputs.xlsx"
wb = load_workbook(file_path)
ws = wb.active

col = 2
row = 3
for outcome in [
    "teachers_uncertified",
    "teachers_secondary_math_outoffield",
    "teachers_secondary_science_outoffield",
    "teachers_secondary_cte_outoffield",
    "class_size_mean_elem",
]:
    te = results.loc[outcome, "te"]
    se = results.loc[outcome, "se"]
    pvalue = (
        scipy.stats.t.sf(np.abs(te / se), n - 1) * 2
    )  # two-sided pvalue = Prob(abs(t)>tt)

    ws.cell(row=row, column=col).value = analysis.coef_with_stars(
        te, pvalue=pvalue, n_tests=1, digits=2
    )
    row = row + 1
    ws.cell(row=row, column=col).value = analysis.format_se(se, 2)
    row = row + 1

wb.save(file_path)
