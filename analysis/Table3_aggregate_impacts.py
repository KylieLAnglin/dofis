# %%

import pandas as pd
import os
from openpyxl import load_workbook

from dofis import start

df = pd.read_excel((os.path.join(start.TABLE_PATH, "aggregated_results.xlsx")))

# %%


def df_to_excel(file, df, df_columns, start_col, start_row, all_ints=True):
    wb = load_workbook(file)

    sheet_name = "ag"
    sheet_to_replace = wb["ag"]
    wb.remove(sheet_to_replace)
    ws = wb.create_sheet("ag")
    # ws = wb[new_sheet]

    col_n = start_col
    for col in df_columns:
        row_n = start_row
        for ob in df[col]:
            ws.cell(row=row_n, column=col_n).value = ob
            row_n = row_n + 1
        col_n = col_n + 1
    wb.save(file)


df_to_excel(
    file=start.TABLE_PATH + "formatted_results/Table3.xlsx",
    df=df,
    df_columns=["outcomes", "tes", "ses"],
    start_col=1,
    start_row=2,
)
