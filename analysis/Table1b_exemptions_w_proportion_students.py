# %%

import pandas as pd
import os
from openpyxl import load_workbook

from dofis import start

# %%


def df_to_excel(file, df, df_columns, start_col, start_row, all_ints=True):
    wb = load_workbook(file)
    ws = wb.active

    col_n = start_col
    for col in df_columns:
        row_n = start_row
        for ob in df[col]:
            ws.cell(row=row_n, column=col_n).value = ob
            row_n = row_n + 1
        col_n = col_n + 1
    wb.save(file)


# %%
data = pd.read_csv(
    (os.path.join(start.DATA_PATH, "clean/", "master_data_district.csv"))
)
data = data[data.year == 2019]
# data = data[data.doi == 1]

data["students_doi"] = data.students_num * data.doi
print(data["students_doi"].sum() / data["students_num"].sum())
# %%


stubnames = sorted(
    set(
        [match[0] for match in data.columns.str.findall(r"reg.*").values if match != []]
    )
)

proportion = []
number = []
proportion_students = []
number_students = []
for reg in stubnames:
    proportion.append(data[reg].mean())
    number.append(data[reg].sum())
    data[reg + "_students"] = data[reg] * data.students_num
    proportion_students.append(
        data[reg + "_students"].sum() / data["students_num"].sum()
    )
    number_students.append(data[reg + "_students"].sum())


regs = pd.DataFrame(
    {
        "law": stubnames,
        "proportion": proportion,
        "number": number,
        "proportion_students": proportion_students,
        "total_students": number_students,
    }
).sort_values(by=["proportion"], ascending=False)
regs.head(10)


# %%


df_to_excel(
    file=os.path.join(start.TABLE_PATH, "Top Exemptions Students.xlsx"),
    df=regs,
    df_columns=["law", "proportion", "number", "proportion_students", "total_students"],
    start_col=4,
    start_row=3,
)
