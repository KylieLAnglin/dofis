# %%
import os

import numpy as np
import pandas as pd
import statsmodels.formula.api as smf
from openpyxl import load_workbook

from dofis import start

# %%

data = pd.read_csv(
    os.path.join(start.DATA_PATH, "clean", "r_data_school_2020_comparison.csv"),
    sep=",",
    low_memory=False,
)

data.sample()


def cluster_sample(df: pd.DataFrame, cluster_var: str):
    clusters = np.random.choice(
        df[cluster_var], size=df[cluster_var].nunique(), replace=True
    )

    # define sample as all members who belong cluster
    cluster_sample = df[df[cluster_var].isin(clusters)]

    return cluster_sample


def cluster_bootstrap(df: pd.DataFrame, cluster_var: str, estimation_func, **kwargs):
    sample = cluster_sample(df=df, cluster_var=cluster_var)
    result = estimation_func(cluster_var=cluster_var, **kwargs)
    return result


def did(
    did_df: pd.DataFrame,
    outcome: str,
    treat_indicator: str,
    post_indicator: str,
    treat_post_indicator: str,
    cluster_var: str,
    quietly: bool = True,
):
    formula = (
        outcome
        + "~ 1 + "
        + treat_indicator
        + " + "
        + post_indicator
        + " + "
        + treat_post_indicator
    )

    mod = smf.ols(formula, did_df)
    res = mod.fit()
    res = mod.fit(cov_type="cluster", cov_kwds={"groups": did_df[cluster_var]})
    if not quietly:
        print(res.summary())

    return res


#!/usr/bin/env python
# coding: utf-8

# %%


def simple_did_df(
    group_var: str,
    group: int,
    time_var: str,
    time: int,
    df: pd.DataFrame,
):
    """Subsets a dataframe so that it only contains two groups
    and two time periods,

    Args:
        group_var (str): column containing group variable
        group (int): value in group column of dataframe to be treated
        time_var (str): column containing time variable
        time (int): value in time column of dataframe to be post
        df (pd.DataFrame): dataframe containing group and time columns

    Returns:
        pd.DataFrame: Untreated rows during one pre period, treat and untreated in one post
        new treatm, post, and treat_post columns
    """
    did_df = df[
        (df[time_var] == time) | (df[time_var] == group - 1)
    ]  # limit to two years
    did_df = did_df[
        (data[group_var] == group) | (data[group_var] < time)
    ]  # limit to two groups

    did_df["treat"] = np.where(did_df[group_var] == group, 1, 0)
    did_df["post"] = np.where(did_df[time_var] == time, 1, 0)
    did_df["treat_post"] = did_df.treat * did_df.post

    return did_df


def dids(
    outcome: str, group_var: str, time_var: str, cluster_var: str, df: pd.DataFrame
):

    groups = np.sort(df[df.group != 0][group_var].unique())
    times = np.sort(df[df[time_var] > df[time_var].min()][time_var].unique())
    group_time_combos = [
        {"group": group, "time": time} for time in times for group in groups
    ]

    times = []
    groups = []
    betas = []
    ses = []
    pvalues = []
    results = {}
    for combo in group_time_combos:
        group = combo["group"]
        time = combo["time"]

        did_df = simple_did_df(
            group_var=group_var,
            group=combo["group"],
            time_var=time_var,
            time=combo["time"],
            df=df,
        )

        formula = outcome + " ~ 1 + treat + post + treat_post"
        mod = smf.ols(formula, did_df)
        res = mod.fit(cov_type="cluster", cov_kwds={"groups": did_df[cluster_var]})
        times.append(time)
        groups.append(group)
        betas.append(res.params["treat_post"])
        ses.append(res.bse["treat_post"])
        pvalues.append(res.pvalues["treat_post"])
        results[
            "Group" + str(int(combo["group"])) + "Time" + str(int(combo["time"]))
        ] = res

        results_df = pd.DataFrame(
            list(zip(times, groups, betas, ses, pvalues)),
            columns=["Time", "Group", "TE", "SE", "P"],
        )

    return results


math_results = dids(
    outcome="math_yr15std",
    group_var="group",
    time_var="year",
    cluster_var="district",
    df=data,
)
math_results

# %%

wb = load_workbook(
    "/Users/kylie/dofis/results/Who Needs Rules/results_math_all_groups_and_times.xlsx"
)
ws = wb.active

col = 2
row = 3

for year in range(2013, 2020):
    print(year)
    print(
        math_results["Group" + str(2017) + "Time" + str(int(year))].params["treat_post"]
    )
    # i = list(
    #     math_results.index[(math_results.Group == 2017) & (math_results.Time == 2017)]
    # )[0]
    # print(math_results.at[i, "TE"])

    # ws.cell(row=row, column=col).value = round(results.df.loc[2017]["TE"])
    # print(round(math_results[(math_results.Time == 2017)]["TE"]))
    # wb.save(file)

# %%
