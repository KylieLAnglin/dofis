# %%

import pandas as pd
import numpy as np
import pandas as pd
from openpyxl import load_workbook
import scipy

from dofis import start
from dofis.analysis.library import analysis

# %%
results_path = start.table_path + "/results_subgroup_raw.xlsx"
results = pd.read_excel(results_path).set_index(["subgroup", "outcome"])
# %%
###
# Demographics
###


file_path = start.table_path + "results_subgroup.xlsx"
wb = load_workbook(file_path)
ws = wb.active

row = 7
col = 2

for group in ["rural", "scores25", "hisp25", "black25"]:
    te = results.loc[(group, "math_yr15std"), "te"]
    se = results.loc[(group, "math_yr15std"), "se"]
    pvalue = (
        scipy.stats.t.sf(np.abs(te / se), n - 1) * 2
    )  # two-sided pvalue = Prob(abs(t)>tt)

    ws.cell(row=row, column=col).value = analysis.coef_with_stars(
        te, pvalue=pvalue, n_tests=8, digits=2
    )
    row = row + 1
    ws.cell(row=row, column=col).value = analysis.format_se(se, 2)
    row = row + 1

row = 16
for group in ["rural", "scores25", "hisp25", "black25"]:
    te = results.loc[(group, "reading_yr15std"), "te"]
    se = results.loc[(group, "reading_yr15std"), "se"]
    pvalue = (
        scipy.stats.t.sf(np.abs(te / se), n - 1) * 2
    )  # two-sided pvalue = Prob(abs(t)>tt)

    ws.cell(row=row, column=col).value = analysis.coef_with_stars(
        te, pvalue=pvalue, n_tests=8, digits=2
    )
    row = row + 1
    ws.cell(row=row, column=col).value = analysis.format_se(se, 2)
    row = row + 1

wb.save(file_path)


row = 7
col = 3

for group in ["urban", "scores100", "hisp100", "black100"]:
    te = results.loc[(group, "math_yr15std"), "te"]
    se = results.loc[(group, "math_yr15std"), "se"]
    pvalue = (
        scipy.stats.t.sf(np.abs(te / se), n - 1) * 2
    )  # two-sided pvalue = Prob(abs(t)>tt)

    ws.cell(row=row, column=col).value = analysis.coef_with_stars(
        te, pvalue=pvalue, n_tests=8, digits=2
    )
    row = row + 1
    ws.cell(row=row, column=col).value = analysis.format_se(se, 2)
    row = row + 1

row = 16
for group in ["urban", "scores100", "hisp100", "black100"]:
    te = results.loc[(group, "reading_yr15std"), "te"]
    se = results.loc[(group, "reading_yr15std"), "se"]
    pvalue = (
        scipy.stats.t.sf(np.abs(te / se), n - 1) * 2
    )  # two-sided pvalue = Prob(abs(t)>tt)

    ws.cell(row=row, column=col).value = analysis.coef_with_stars(
        te, pvalue=pvalue, n_tests=8, digits=2
    )
    row = row + 1
    ws.cell(row=row, column=col).value = analysis.format_se(se, 2)
    row = row + 1

wb.save(file_path)

# %%
###
# Exemptions
###

results_path = start.table_path + "/results_exemptions_raw.xlsx"
results = pd.read_excel(results_path).set_index(["subgroup", "outcome", "exempt"])


file_path = start.table_path + "results_subgroup_exemptions_math.xlsx"
wb = load_workbook(file_path)
ws = wb.active

row = 5
for group in [
    "exempt_firstday",
    "exempt_minutes",
    "exempt_lastday",
    "exempt_certification",
    "exempt_probation",
    "exempt_servicedays",
    "exempt_eval",
    "exempt_classsize",
    "exempt_behavior",
]:
    te = results.loc[(group, "math_yr15std", 0), "te"]
    se = results.loc[(group, "math_yr15std", 0), "se"]
    pvalue = (
        scipy.stats.t.sf(np.abs(te / se), n - 1) * 2
    )  # two-sided pvalue = Prob(abs(t)>tt)

    ws.cell(row=row, column=2).value = analysis.coef_with_stars(
        te, pvalue=pvalue, n_tests=18, digits=2
    )
    row = row + 1
    ws.cell(row=row, column=2).value = analysis.format_se(se, 2)
    row = row + 1

row = 5
for group in [
    "exempt_firstday",
    "exempt_minutes",
    "exempt_lastday",
    "exempt_certification",
    "exempt_probation",
    "exempt_servicedays",
    "exempt_eval",
    "exempt_classsize",
    "exempt_behavior",
]:
    te = results.loc[(group, "math_yr15std", 1), "te"]
    se = results.loc[(group, "math_yr15std", 1), "se"]
    pvalue = (
        scipy.stats.t.sf(np.abs(te / se), n - 1) * 2
    )  # two-sided pvalue = Prob(abs(t)>tt)

    ws.cell(row=row, column=3).value = analysis.coef_with_stars(
        te, pvalue=pvalue, n_tests=18, digits=2
    )
    row = row + 1
    ws.cell(row=row, column=3).value = analysis.format_se(se, 2)
    row = row + 1

wb.save(file_path)


# %% Reading
results_path = start.table_path + "/results_exemptions_raw.xlsx"
results = pd.read_excel(results_path).set_index(["subgroup", "outcome", "exempt"])


file_path = start.table_path + "results_subgroup_exemptions_reading.xlsx"
wb = load_workbook(file_path)
ws = wb.active

row = 5
for group in [
    "exempt_firstday",
    "exempt_minutes",
    "exempt_lastday",
    "exempt_certification",
    "exempt_probation",
    "exempt_servicedays",
    "exempt_eval",
    "exempt_classsize",
    "exempt_behavior",
]:
    te = results.loc[(group, "reading_yr15std", 0), "te"]
    se = results.loc[(group, "reading_yr15std", 0), "se"]
    pvalue = (
        scipy.stats.t.sf(np.abs(te / se), n - 1) * 2
    )  # two-sided pvalue = Prob(abs(t)>tt)

    ws.cell(row=row, column=2).value = analysis.coef_with_stars(
        te, pvalue=pvalue, n_tests=18, digits=2
    )
    row = row + 1
    ws.cell(row=row, column=2).value = analysis.format_se(se, 2)
    row = row + 1

row = 5
for group in [
    "exempt_firstday",
    "exempt_minutes",
    "exempt_lastday",
    "exempt_certification",
    "exempt_probation",
    "exempt_servicedays",
    "exempt_eval",
    "exempt_classsize",
    "exempt_behavior",
]:
    te = results.loc[(group, "reading_yr15std", 1), "te"]
    se = results.loc[(group, "reading_yr15std", 1), "se"]
    pvalue = (
        scipy.stats.t.sf(np.abs(te / se), n - 1) * 2
    )  # two-sided pvalue = Prob(abs(t)>tt)

    ws.cell(row=row, column=3).value = analysis.coef_with_stars(
        te, pvalue=pvalue, n_tests=18, digits=2
    )
    row = row + 1
    ws.cell(row=row, column=3).value = analysis.format_se(se, 2)
    row = row + 1

wb.save(file_path)
