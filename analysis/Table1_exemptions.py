# %%

import pandas as pd
import os
from openpyxl import load_workbook

from dofis import start

data = pd.read_csv(
    (os.path.join(start.DATA_PATH, "clean/", "master_data_district.csv"))
)
data = data[data.year == 2019]
data = data[data.doi == 1]


def df_to_excel(file, df, df_columns, start_col, start_row, all_ints=True):
    wb = load_workbook(file)
    ws = wb.active

    col_n = start_col
    for col in df_columns:
        row_n = start_row
        for ob in df[col]:
            ws.cell(row=row_n, column=col_n).value = ob
            row_n = row_n + 1
        col_n = col_n + 1
    wb.save(file)


# %%

stubnames = sorted(
    set(
        [match[0] for match in data.columns.str.findall(r"reg.*").values if match != []]
    )
)

proportion = []
number = []
for reg in stubnames:
    proportion.append(data[reg].mean())
    number.append(data[reg].sum())

regs = pd.DataFrame(
    {"law": stubnames, "proportion": proportion, "number": number}
).sort_values(by=["proportion"], ascending=False)
regs.head(10)


# %%


df_to_excel(
    file=os.path.join(start.TABLE_PATH, "Top Exemptions.xlsx"),
    df=regs,
    df_columns=["law", "proportion"],
    start_col=4,
    start_row=3,
)


df_to_excel(
    file=os.path.join(start.TABLE_PATH, "formatted_results/Table1a.xlsx"),
    df=regs,
    df_columns=["law", "proportion"],
    start_col=4,
    start_row=3,
)
