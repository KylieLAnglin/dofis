# %%

import numpy as np
import pandas as pd
import statsmodels.formula.api as smf
import scipy
from openpyxl import load_workbook


from dofis.analysis.library import analysis, characteristics, regulations, tables
from dofis import start

# %%
df = pd.read_csv(start.DATA_PATH + "clean/master_data_school.csv")
df["teachers_new"] = df.teachers_new_num / df.teachers

# %%

# %%
FILE = start.TABLE_PATH + "Simple Pre-post Differences.xlsx"
wb = load_workbook(FILE)
ws = wb["Inputs"]

# %%
GROUPS_FOR_TABLE = ["2017", "2018", "2019", "2020+", "opt-out"]


def simple_pre_post_table(
    start_row: int, start_col: int, outcome: str, digits: int = 2
):
    row = start_row
    col = start_col
    for group in GROUPS_FOR_TABLE:
        col = start_col
        for year in [2016, 2017, 2018, 2019, 2020, 2021]:
            value = df[(df.group == group) & (df.year == year)][outcome].mean()
            ws.cell(row=row, column=col).value = round(value, digits)
            col = col + 1
            print(year, group, str(value))
        row = row + 1

    wb.save(FILE)


simple_pre_post_table(
    start_row=3, start_col=2, outcome="teachers_uncertified", digits=3
)
simple_pre_post_table(start_row=11, start_col=2, outcome="teachers_new_num", digits=1)
simple_pre_post_table(start_row=19, start_col=2, outcome="teachers", digits=1)
simple_pre_post_table(
    start_row=27, start_col=2, outcome="teachers_tenure_ave", digits=1
)
simple_pre_post_table(start_row=35, start_col=2, outcome="class_size_k", digits=1)
simple_pre_post_table(start_row=43, start_col=2, outcome="class_size_3", digits=1)
simple_pre_post_table(start_row=51, start_col=2, outcome="class_size_5", digits=1)
simple_pre_post_table(
    start_row=59, start_col=2, outcome="class_size_sec_lang", digits=1
)
simple_pre_post_table(
    start_row=67, start_col=2, outcome="class_size_sec_math", digits=1
)
simple_pre_post_table(
    start_row=75, start_col=2, outcome="teachers_secondary_math_outoffield", digits=2
)
simple_pre_post_table(
    start_row=83, start_col=2, outcome="teachers_secondary_science_outoffield", digits=2
)

simple_pre_post_table(start_row=91, start_col=2, outcome="stu_teach_ratio", digits=2)
simple_pre_post_table(start_row=99, start_col=2, outcome="teachers_nodegree", digits=2)
simple_pre_post_table(start_row=107, start_col=2, outcome="teachers_msdegree", digits=2)
simple_pre_post_table(
    start_row=115, start_col=2, outcome="teacher_salary_ave", digits=0
)

# %%
FILE = start.TABLE_PATH + "Simple Pre-post Differences.xlsx"
wb = load_workbook(FILE)
ws = wb["Outputs"]


simple_pre_post_table(start_row=3, start_col=2, outcome="elem_math_yr15std", digits=2)
simple_pre_post_table(
    start_row=11, start_col=2, outcome="elem_reading_yr15std", digits=2
)
simple_pre_post_table(
    start_row=19, start_col=2, outcome="middle_math_yr15std", digits=2
)
simple_pre_post_table(
    start_row=27, start_col=2, outcome="middle_reading_yr15std", digits=2
)
simple_pre_post_table(
    start_row=35, start_col=2, outcome="middle_science_yr15std", digits=2
)
simple_pre_post_table(start_row=43, start_col=2, outcome="algebra_yr15std", digits=2)
simple_pre_post_table(start_row=51, start_col=2, outcome="biology_yr15std", digits=2)

# %%
FILE = start.TABLE_PATH + "Simple Pre-post Differences.xlsx"
wb = load_workbook(FILE)
ws = wb["Other"]


simple_pre_post_table(start_row=3, start_col=2, outcome="perf_attendance", digits=2)
simple_pre_post_table(start_row=11, start_col=2, outcome="perf_hsdrop", digits=2)
simple_pre_post_table(start_row=19, start_col=2, outcome="perf_stuattend", digits=2)
simple_pre_post_table(start_row=27, start_col=2, outcome="perf_studays", digits=2)

# %%
