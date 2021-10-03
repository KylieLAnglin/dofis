#!/usr/bin/env python
# coding: utf-8

# %%
import os
import sys
import random

import numpy as np
import pandas as pd
from openpyxl import load_workbook

from dofis import start
from dofis.analysis.library import analysis


# %%

data = pd.read_csv(
    os.path.join(start.DATA_PATH, "clean", "r_data_school_2020_comparison.csv"),
    sep=",",
    low_memory=False,
)

data.sample()


def print_results_for_group(results: dict, group: int, col: int, start_row: int):
    col = col
    row = start_row

    for year in range(2013, 2020):
        te = results[group][year].params["treat_post"]
        pvalue = results[group][year].pvalues["treat_post"]
        ws.cell(row=row, column=col).value = analysis.coef_with_stars(
            te, pvalue=pvalue, n_tests=1, digits=2
        )
        row = row + 1
        se = results[group][year].bse["treat_post"]
        ws.cell(row=row, column=col).value = analysis.format_se(se, 2)
        row = row + 1
        wb.save(file_path)


# %% Math
math_results = analysis.dids(
    outcome="math",
    group_var="group",
    time_var="year",
    cluster_var="district",
    df=data,
)


def simple_weighted_average(results: dict, data: pd.DataFrame):
    """Calculate simple weighted average across cohorts and years

    Args:
        results (dict): dictionary result from analysis.dids
    """
    post_results = []
    post_results.append(results[2017][2017].params["treat_post"])
    post_results.append(results[2017][2018].params["treat_post"])
    post_results.append(results[2017][2019].params["treat_post"])
    post_results.append(results[2018][2018].params["treat_post"])
    post_results.append(results[2018][2019].params["treat_post"])
    post_results.append(results[2019][2019].params["treat_post"])

    samples = []
    samples.append(len(data[(data.group == 2017) & (data.year == 2017)]))
    samples.append(len(data[(data.group == 2017) & (data.year == 2018)]))
    samples.append(len(data[(data.group == 2017) & (data.year == 2019)]))
    samples.append(len(data[(data.group == 2018) & (data.year == 2018)]))
    samples.append(len(data[(data.group == 2018) & (data.year == 2019)]))
    samples.append(len(data[(data.group == 2019) & (data.year == 2019)]))

    total = sum(samples)

    weights = [sample / total for sample in samples]
    weighted_results = [
        result * weight for result, weight in zip(post_results, weights)
    ]

    weighted_result = sum(weighted_results)

    return weighted_result


math_simple = simple_weighted_average(math_results, data)

# math_simples = []
# for i in range(0, 1000):
#     sample_df = data.sample(len(data), replace=True)
#     math_results_boot = analysis.dids(
#         outcome="math_yr15std",
#         group_var="group",
#         time_var="year",
#         cluster_var="district",
#         df=sample_df,
#         verbose=False,
#     )
#     result = simple_weighted_average(math_results_boot, sample_df)
#     math_simples.append(result)

# tes = te.append()
# for group in [2017, 2018, 2019]:
# for year in range(2013, 2020):


# %%

file_path = start.TABLE_PATH + "results_math_all_groups_and_times.xlsx"
wb = load_workbook(file_path)
ws = wb.active

print_results_for_group(results=math_results, group=2017, col=2, start_row=3)
print_results_for_group(results=math_results, group=2018, col=3, start_row=3)
print_results_for_group(results=math_results, group=2019, col=4, start_row=3)

wb.save(file_path)


# %% Reading
reading_results = analysis.dids(
    outcome="reading",
    group_var="group",
    time_var="year",
    cluster_var="district",
    df=data,
)

file_path = start.TABLE_PATH + "results_reading_all_groups_and_times.xlsx"
wb = load_workbook(file_path)
ws = wb.active

print_results_for_group(results=reading_results, group=2017, col=2, start_row=3)
print_results_for_group(results=reading_results, group=2018, col=3, start_row=3)
print_results_for_group(results=reading_results, group=2019, col=4, start_row=3)

wb.save(file_path)

# %%
