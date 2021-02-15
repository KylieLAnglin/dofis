#!/usr/bin/env python
# coding: utf-8

# %%
import os
import sys

import matplotlib.mlab as mlab
import matplotlib.pyplot as plt
import numpy as np
import pandas as pd
import statsmodels.formula.api as smf
from linearmodels import PanelOLS
from openpyxl import load_workbook
from patsy import dmatrices

from dofis.analysis.library import start
from dofis.analysis.library import analysis


# %%

data = pd.read_csv(
    os.path.join(start.data_path, "clean", "gdid_school.csv"),
    sep=",",
    low_memory=False,
)
# data = data[data.analytic_sample]
data = data[data.doi_year > 2016]
data = data[data.doi_year <= 2020]
print(data.doi_year.value_counts())

data.sample()

# %%

# convert year to datetime
df = data.reset_index()
df["year_index"] = pd.to_datetime(df["year"], format="%Y")
df["campus_index"] = df.campus
df = df.set_index(["campus_index", "year_index"])
df[
    [
        "campus",
        "year",
        "doi_year",
        "treatpost",
        "yearpost",
        "post1",
        "math",
        "reading",
        "elem_math",
    ]
].sample(5)


# %% Get table ready

GDID_MODEL = " ~ + 1 + treatpost + C(year) + EntityEffects"
LINEAR_GDID_MODEL = (
    " ~ + 1 + treatpost + yearpost + " "yearpre + C(year) + EntityEffects"
)
EVENT_STUDY_MODEL = (
    " ~ + 1 + pre5 + pre4 + pre3 + pre2 + "
    "post1 + post2 + post3 + C(year)+ EntityEffects"
)

# %%
mod = PanelOLS.from_formula("math" + EVENT_STUDY_MODEL, df)
res = mod.fit(cov_type="clustered", clusters=df.district)
print(res)

# %%


def results_table(data: pd.DataFrame, outcome: str, file_name: str):
    """Generate results table for three specifications

    Args:
        data (pd.DataFrame): Dataframe containing correct sample
        outcome (str): outcome to estimate
        file_name (str): name of file for table
    """
    file = start.table_path + file_name
    wb = load_workbook(file)
    ws = wb.active

    # GDID
    mod = PanelOLS.from_formula(outcome + GDID_MODEL, data)
    res = mod.fit(cov_type="clustered", clusters=data.district)
    # res = mod.fit(cov_type='clustered', cluster_entity = True,
    # cluster_time = True)
    print(res)
    ws.cell(row=3, column=2).value = analysis.coef_with_stars(
        res.params["treatpost[T.True]"], res.pvalues["treatpost[T.True]"]
    )
    ws.cell(row=4, column=2).value = analysis.format_se(
        res.std_errors["treatpost[T.True]"]
    )

    # Linear GDID
    mod = PanelOLS.from_formula(outcome + LINEAR_GDID_MODEL, data)
    res = mod.fit(cov_type="clustered", clusters=data.district)
    print(res)

    ws.cell(row=6, column=2).value = analysis.coef_with_stars(
        res.params["yearpre"], res.pvalues["yearpre"]
    )
    ws.cell(row=7, column=2).value = analysis.format_se(res.std_errors["yearpre"])

    ws.cell(row=8, column=2).value = analysis.coef_with_stars(
        res.params["treatpost[T.True]"], res.pvalues["treatpost[T.True]"]
    )
    ws.cell(row=9, column=2).value = analysis.format_se(
        res.std_errors["treatpost[T.True]"]
    )

    ws.cell(row=10, column=2).value = analysis.coef_with_stars(
        res.params["yearpost"], res.pvalues["yearpost"]
    )
    ws.cell(row=11, column=2).value = analysis.format_se(res.std_errors["yearpost"])

    # Event Study
    mod = PanelOLS.from_formula(outcome + EVENT_STUDY_MODEL, data)
    res = mod.fit(cov_type="clustered", clusters=data.district)
    print(res)
    row = 3
    for coef in ["pre5", "pre4", "pre3", "pre2", "post1", "post2", "post3"]:
        ws.cell(row=row, column=4).value = analysis.coef_with_stars(
            res.params[coef], res.pvalues[coef]
        )
        row = row + 1
        ws.cell(row=row, column=4).value = analysis.format_se(res.std_errors[coef])
        row = row + 1

    wb.save(file)


# %% Math


results_table(data=df, outcome="math", file_name="gdid_and_event_math_school.xlsx")

results_table(
    data=df, outcome="reading", file_name="gdid_and_event_reading_school.xlsx"
)

# %% Event Study Graphs - Math

mod = PanelOLS.from_formula("math" + EVENT_STUDY_MODEL, df)
res = mod.fit(cov_type="clustered", clusters=df.district)

nonparametric = []
nonparametric_se = []
for coef in ["pre5", "pre4", "pre3", "pre2", "pre1", "post1", "post2", "post3"]:
    nonpar = 0
    nonpar_se = 0
    if coef != "pre1":
        nonpar = res.params[coef]
        nonpar_se = res.std_errors[coef]
    nonparametric.append(nonpar)
    nonparametric_se.append(nonpar_se)
coef_df = pd.DataFrame(
    {
        "coef": nonparametric,
        "err": nonparametric_se,
        "year": [-5, -4, -3, -2, -1, 1, 2, 3],
    }
)
coef_df["lb"] = coef_df.coef - (1.96 * coef_df.err)
coef_df["ub"] = coef_df.coef + (1.96 * coef_df.err)
coef_df["errsig"] = coef_df.err * 1.96

fig, ax = plt.subplots(figsize=(8, 5))

coef_df.plot(
    x="year", y="coef", kind="bar", ax=ax, color="none", yerr="errsig", legend=False
)
ax.set_ylabel("")
ax.set_xlabel("")
ax.scatter(
    x=pd.np.arange(coef_df.shape[0]),
    marker="s",
    s=120,
    y=coef_df["coef"],
    color="black",
)
ax.axhline(y=0, linestyle="--", color="black", linewidth=4)
ax.xaxis.set_ticks_position("none")
_ = ax.set_xticklabels(
    ["Pre5", "Pre4", "Pre3", "Pre2", "Pre1", "Post1", "Post2", "Post3"], rotation=0
)
# ax.set_title('Impact on Student Achievement - Event Study Coefficients',
# fontsize = 16)

fig.savefig(start.table_path + "math_event_study" + ".png", bbox_inches="tight")


# %% Event Study Graphs - Reading

mod = PanelOLS.from_formula(EVENT_STUDY_MODEL, df[df.reading_test == 1])
res = mod.fit(cov_type="clustered", clusters=df[df.reading_test == 1].district)
nonparametric = []
nonparametric_se = []
for coef in ["pre5", "pre4", "pre3", "pre2", "pre1", "post1", "post2", "post3"]:
    nonpar = 0
    nonpar_se = 0
    if coef != "pre1":
        nonpar = res.params[coef]
        nonpar_se = res.std_errors[coef]
    nonparametric.append(nonpar)
    nonparametric_se.append(nonpar_se)
coef_df = pd.DataFrame(
    {
        "coef": nonparametric,
        "err": nonparametric_se,
        "year": [-5, -4, -3, -2, -1, 1, 2, 3],
    }
)
coef_df["lb"] = coef_df.coef - (1.96 * coef_df.err)
coef_df["ub"] = coef_df.coef + (1.96 * coef_df.err)
coef_df["errsig"] = coef_df.err * 1.96

fig, ax = plt.subplots(figsize=(8, 5))

coef_df.plot(
    x="year", y="coef", kind="bar", ax=ax, color="none", yerr="errsig", legend=False
)
ax.set_ylabel("")
ax.set_xlabel("")
ax.scatter(
    x=pd.np.arange(coef_df.shape[0]),
    marker="s",
    s=120,
    y=coef_df["coef"],
    color="black",
)
ax.axhline(y=0, linestyle="--", color="black", linewidth=4)
ax.xaxis.set_ticks_position("none")
_ = ax.set_xticklabels(
    ["Pre5", "Pre4", "Pre3", "Pre2", "Pre1", "Post1", "Post2", "Post3"], rotation=0
)
# ax.set_title('Impact on Student Achievement - Event Study Coefficients',
#  fontsize = 16)

fig.savefig(start.table_path + "reading_event_study" + ".png", bbox_inches="tight")
