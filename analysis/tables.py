from openpyxl import Workbook
from openpyxl import load_workbook

def df_to_excel(file, df, df_columns, start_col, start_row):
    wb = load_workbook(file)
    ws = wb.active

    col_n = start_col
    for col in df_columns:
        row_n = start_row
        for ob in df[col]:
            if ob < .01:
                ob = '<.01'
            ws.cell(row=row_n, column=col_n).value = ob
            row_n = row_n + 1
        col_n = col_n + 1
    wb.save(file)