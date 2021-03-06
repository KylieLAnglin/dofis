# %%
import pandas as pd
import statsmodels.formula.api as smf
from openpyxl import load_workbook


from dofis.analysis.library import start
from dofis.analysis.library import characteristics
from dofis.analysis.library import analysis


INPUTS = [
    "class_size_k",
    "class_size_1",
    "class_size_2",
    "class_size_3",
    "class_size_4",
    "class_size_5",
    "perf_attendance",
    "perf_stuattend",
    "perf_studays",
    "stu_teach_ratio",
    "days_max",
    "days_mean",
    "days_min",
    "teachers_certified",
    "teachers_uncertified",
    "teachers_secondary_math_certified",
    "teachers_secondary_math_uncertified",
    "teachers_secondary_math_outoffield",
    "teachers_secondary_science_outoffield",
    "teachers_secondary_science_uncertified",
    "teachers_secondary_cte_uncertified",
    "teachers_secondary_cte_outoffield",
]


COLUMNS = (
    ["year", "doi_year", "campus", "campname", "district", "distname"]
    + characteristics.POTENTIAL_COVARIATES
    + INPUTS
    + [
        "exempt_minutes",
        "exempt_certification",
        "exempt_classsize",
        "exempt_servicedays",
    ]
)

# %%
data_school = pd.read_csv(
    os.path.join(start.data_path, "clean", "master_data_school.csv"),
    sep=",",
    low_memory=False,
)
data_school = data_school[data_school.doi == 1]
data_school = data_school[COLUMNS]
data_school["const"] = 1

data_school = data_school.set_index("campus")


def input(outcome: str, data: pd.DataFrame, table: str, col: int, start_row: int):

    file = start.table_path + table
    wb = load_workbook(file)
    ws = wb.active

    # Estimates
    ws.cell(row=start_row, column=col).value = round(data[outcome].mean(), 4)
    sd_row = start_row + 1
    ws.cell(row=sd_row, column=col).value = (
        "[" + str(round(data[outcome].std(), 4)) + "]"
    )

    wb.save(file)


def print_columns(data, col):
    row = 4
    for var in [
        "teachers_uncertified",
        "teachers_secondary_cte_uncertified",
        "teachers_secondary_math_uncertified",
        "teachers_secondary_science_uncertified",
        "teachers_secondary_math_outoffield",
        "teachers_secondary_science_outoffield",
    ]:
        input(
            outcome=var,
            data=data,
            table="balance_exemptions_certification.xlsx",
            col=col,
            start_row=row,
        )
        row = row + 2


print_columns(data_school[(data_school.year == 2016)], 2)
print_columns(data_school[(data_school.year == 2017)], 3)
print_columns(
    data_school[(data_school.year == 2017) & (data_school.exempt_certification == 1)], 4
)
print_columns(
    data_school[(data_school.year == 2017) & (data_school.exempt_certification == 0)], 5
)

# %%
row = 4
for var in [
    "teachers_uncertified",
    "teachers_secondary_cte_uncertified",
    "teachers_secondary_math_uncertified",
    "teachers_secondary_science_uncertified",
    "teachers_secondary_math_outoffield",
    "teachers_secondary_science_outoffield",
]:
    input(
        outcome=var,
        data=data_school[(data_school.year == 2016)],
        table="balance_exemptions_certification.xlsx",
        col=2,
        start_row=row,
    )
    row = row + 2

row = 4
for var in [
    "teachers_uncertified",
    "teachers_secondary_cte_uncertified",
    "teachers_secondary_math_uncertified",
    "teachers_secondary_science_uncertified",
    "teachers_secondary_math_outoffield",
    "teachers_secondary_science_outoffield",
]:
    input(
        outcome=var,
        data=data_school[(data_school.year == 2017)],
        table="balance_exemptions_certification.xlsx",
        col=3,
        start_row=row,
    )
    row = row + 2

row = 4
for var in [
    "teachers_uncertified",
    "teachers_secondary_cte_uncertified",
    "teachers_secondary_math_uncertified",
    "teachers_secondary_science_uncertified",
    "teachers_secondary_math_outoffield",
    "teachers_secondary_science_outoffield",
]:
    input(
        outcome=var,
        data=data_school[
            (data_school.year == 2017) & (data_school.exempt_certification == 1)
        ],
        table="balance_exemptions_certification.xlsx",
        col=4,
        start_row=row,
    )
    row = row + 2

for var in [
    "teachers_uncertified",
    "teachers_secondary_cte_uncertified",
    "teachers_secondary_math_uncertified",
    "teachers_secondary_science_uncertified",
    "teachers_secondary_math_outoffield",
    "teachers_secondary_science_outoffield",
]:
    input(
        outcome=var,
        data=data_school[
            (data_school.year == 2017) & (data_school.exempt_certification == 0)
        ],
        table="balance_exemptions_certification.xlsx",
        col=5,
        start_row=row,
    )
    row = row + 2
