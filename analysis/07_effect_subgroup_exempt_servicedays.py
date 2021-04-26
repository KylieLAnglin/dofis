# %%

import pandas as pd
import numpy as np
import pandas as pd
from openpyxl import load_workbook
import scipy

from dofis.analysis.library import start
from dofis.analysis.library import analysis

# %%
results_path = start.table_path + "/exempt_servicedays_disag_math.xlsx"
results = pd.read_excel(results_path)

### Table 5
math_disag = math_disag.rename(
    columns={
        "disag.math.group": "group",
        "disag.math.t": "year",
        "disag.math.att": "te",
        "disag.math.se": "se",
    }
)

file_path = start.table_path + "results_math.xlsx"
wb = load_workbook(file_path)
ws = wb.active

row = 3
col = 2
for year in [2013, 2014, 2015, 2016, 2017, 2018, 2019]:
    te = float(
        math_disag.loc[(math_disag.year == year) & (math_disag.group == 2017), "te"]
    )
    se = float(
        math_disag.loc[(math_disag.year == year) & (math_disag.group == 2017), "se"]
    )
    pvalue = (
        scipy.stats.t.sf(np.abs(te / se), n - 1) * 2
    )  # two-sided pvalue = Prob(abs(t)>tt)
    ws.cell(row=row, column=col).value = analysis.coef_with_stars(
        te, pvalue=pvalue, n_tests=1, digits=2
    )
    row = row + 1
    ws.cell(row=row, column=col).value = analysis.format_se(se, 2)
    row = row + 1


col = 3
row = 3
for year in [2013, 2014, 2015, 2016, 2017, 2018, 2019]:
    te = float(
        math_disag.loc[(math_disag.year == year) & (math_disag.group == 2018), "te"]
    )
    se = float(
        math_disag.loc[(math_disag.year == year) & (math_disag.group == 2018), "se"]
    )
    pvalue = (
        scipy.stats.t.sf(np.abs(te / se), n - 1) * 2
    )  # two-sided pvalue = Prob(abs(t)>tt)
    ws.cell(row=row, column=col).value = analysis.coef_with_stars(
        te, pvalue=pvalue, n_tests=1, digits=2
    )
    row = row + 1
    ws.cell(row=row, column=col).value = analysis.format_se(se, 2)
    row = row + 1

col = 4
row = 3
for year in [2013, 2014, 2015, 2016, 2017, 2018, 2019]:
    te = float(
        math_disag.loc[(math_disag.year == year) & (math_disag.group == 2019), "te"]
    )
    se = float(
        math_disag.loc[(math_disag.year == year) & (math_disag.group == 2019), "se"]
    )
    pvalue = (
        scipy.stats.t.sf(np.abs(te / se), n - 1) * 2
    )  # two-sided pvalue = Prob(abs(t)>tt)
    ws.cell(row=row, column=col).value = analysis.coef_with_stars(
        te, pvalue=pvalue, n_tests=1, digits=2
    )
    row = row + 1
    ws.cell(row=row, column=col).value = analysis.format_se(se, 2)
    row = row + 1

wb.save(file_path)


###
# Reading
###
reading_disag = reading_disag.rename(
    columns={
        "disag.reading.group": "group",
        "disag.reading.t": "year",
        "disag.reading.att": "te",
        "disag.reading.se": "se",
    }
)

results_path = start.table_path + "/exempt_servicedays_disag_math.xlsx"
results = pd.read_excel(results_path).set_index(["subgroup", "outcome"])

wb = load_workbook(file_path)
ws = wb.active

row = 3
col = 2
for year in [2013, 2014, 2015, 2016, 2017, 2018, 2019]:
    te = float(
        reading_disag.loc[
            (reading_disag.year == year) & (reading_disag.group == 2017), "te"
        ]
    )
    se = float(
        reading_disag.loc[
            (reading_disag.year == year) & (reading_disag.group == 2017), "se"
        ]
    )
    pvalue = (
        scipy.stats.t.sf(np.abs(te / se), n - 1) * 2
    )  # two-sided pvalue = Prob(abs(t)>tt)
    ws.cell(row=row, column=col).value = analysis.coef_with_stars(
        te, pvalue=pvalue, n_tests=7, digits=2
    )
    row = row + 1
    ws.cell(row=row, column=col).value = analysis.format_se(se, 2)
    row = row + 1


col = 3
row = 3
for year in [2013, 2014, 2015, 2016, 2017, 2018, 2019]:
    te = float(
        reading_disag.loc[
            (reading_disag.year == year) & (reading_disag.group == 2018), "te"
        ]
    )
    se = float(
        reading_disag.loc[
            (reading_disag.year == year) & (reading_disag.group == 2018), "se"
        ]
    )
    pvalue = (
        scipy.stats.t.sf(np.abs(te / se), n - 1) * 2
    )  # two-sided pvalue = Prob(abs(t)>tt)
    ws.cell(row=row, column=col).value = analysis.coef_with_stars(
        te, pvalue=pvalue, n_tests=7, digits=2
    )
    row = row + 1
    ws.cell(row=row, column=col).value = analysis.format_se(se, 2)
    row = row + 1


col = 4
row = 3
for year in [2013, 2014, 2015, 2016, 2017, 2018, 2019]:
    te = float(
        reading_disag.loc[
            (reading_disag.year == year) & (reading_disag.group == 2019), "te"
        ]
    )
    se = float(
        reading_disag.loc[
            (reading_disag.year == year) & (reading_disag.group == 2019), "se"
        ]
    )
    pvalue = (
        scipy.stats.t.sf(np.abs(te / se), n - 1) * 2
    )  # two-sided pvalue = Prob(abs(t)>tt)
    ws.cell(row=row, column=col).value = analysis.coef_with_stars(
        te, pvalue=pvalue, n_tests=7, digits=2
    )
    row = row + 1
    ws.cell(row=row, column=col).value = analysis.format_se(se, 2)
    row = row + 1


wb.save(file_path)