# %%
import pandas as pd
import numpy as np
import pandas as pd
from openpyxl import load_workbook
import scipy

from dofis.analysis.library import start
from dofis.analysis.library import analysis

MATH_AGG = start.table_path + "results_math_ag_raw_std.xlsx"
math_agg = pd.read_excel(MATH_AGG)

MATH_DISAG = start.table_path + "results_math_disag_raw_std.xlsx"
math_disag = pd.read_excel(MATH_DISAG)

READING_AGG = start.table_path + "results_reading_ag_raw_std.xlsx"
reading_agg = pd.read_excel(READING_AGG)

READING_DISAG = start.table_path + "results_reading_disag_raw_std.xlsx"
reading_disagg = pd.read_excel(READING_DISAG)

data = pd.read_csv(start.data_path + "clean/r_data_school_2020_comparison.csv")
n = data.district.nunique()


# %% Tables 4 & 5 Main


file_path = start.table_path + "results_main_standardize_within_year.xlsx"
wb = load_workbook(file_path)
ws = wb.active

###
# Math
###

math_agg = math_agg.rename(
    columns={
        "agg.simple.math.overall.att": "overall_te",
        "agg.simple.math.overall.se": "overall_se",
        "agg.dynamic.math.egt": "year",
        "agg.dynamic.math.att.egt": "te",
        "agg.dynamic.math.se.egt": "se",
    }
)

# Overall
te = math_agg.loc[0, "overall_te"]
se = math_agg.loc[0, "overall_se"]
pvalue = (
    scipy.stats.t.sf(np.abs(te / se), n - 1) * 2
)  # two-sided pvalue = Prob(abs(t)>tt)

ws.cell(row=3, column=2).value = analysis.coef_with_stars(
    te, pvalue=pvalue, n_tests=1, digits=2
)
ws.cell(row=4, column=2).value = analysis.format_se(se, 2)


# Dynamic
row = 4
col = 3
for year in [0, 1, 2]:
    te = float(math_agg.loc[math_agg.year == year, "te"])
    se = float(math_agg.loc[math_agg.year == year, "se"])
    pvalue = (
        scipy.stats.t.sf(np.abs(te / se), n - 1) * 2
    )  # two-sided pvalue = Prob(abs(t)>tt)
    ws.cell(row=row, column=col).value = analysis.coef_with_stars(
        te, pvalue=pvalue, n_tests=3, digits=2
    )
    row = row + 1
    ws.cell(row=row, column=col).value = analysis.format_se(se, 2)
    row = row + 2

wb.save(file_path)


###
# Reading
###

reading_agg = reading_agg.rename(
    columns={
        "agg.simple.reading.overall.att": "overall_te",
        "agg.simple.reading.overall.se": "overall_se",
        "agg.dynamic.reading.egt": "year",
        "agg.dynamic.reading.att.egt": "te",
        "agg.dynamic.reading.se.egt": "se",
    }
)


# Overall
te = reading_agg.loc[0, "overall_te"]
se = reading_agg.loc[0, "overall_se"]
pvalue = (
    scipy.stats.t.sf(np.abs(te / se), n - 1) * 2
)  # two-sided pvalue = Prob(abs(t)>tt)

ws.cell(row=13, column=2).value = analysis.coef_with_stars(
    te, pvalue=pvalue, n_tests=3, digits=2
)
ws.cell(row=14, column=2).value = analysis.format_se(se, 2)


# Dynamic
row = 14
col = 3
for year in [0, 1, 2]:
    te = float(reading_agg.loc[reading_agg.year == year, "te"])
    se = float(reading_agg.loc[reading_agg.year == year, "se"])
    pvalue = (
        scipy.stats.t.sf(np.abs(te / se), n - 1) * 2
    )  # two-sided pvalue = Prob(abs(t)>tt)
    ws.cell(row=row, column=col).value = analysis.coef_with_stars(
        te, pvalue=pvalue, n_tests=1, digits=2
    )
    row = row + 1
    ws.cell(row=row, column=col).value = analysis.format_se(se, 2)
    row = row + 2

wb.save(file_path)


# %%
