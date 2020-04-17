from openpyxl import load_workbook
import statsmodels.formula.api as smf

def df_to_excel(file, df, df_columns, start_col, start_row, all_ints = True):
    wb = load_workbook(file)
    ws = wb.active

    col_n = start_col
    for col in df_columns:
        row_n = start_row
        for ob in df[col]:
            if all_ints:
                try: 
                    if ob < .01:
                        ob = '<.01'
                except:
                    ob = ob
            ws.cell(row=row_n, column=col_n).value = ob
            row_n = row_n + 1
        col_n = col_n + 1
    wb.save(file)


def model_to_excel(data, y, x_list, file, start_row, start_col):
    wb = load_workbook(file)
    ws = wb.active

    col_n = start_col
    row_n = start_row

    # create formula
    i = 1
    x_str = ''
    for x in x_list:
        if i == 1:
            x_str = x_str + x
        if i > 1:
            x_str = x_str + ' + ' + x
        i += 1
    x_str = x_str + '- 1'
    formula = y + ' ~ ' + x_str

    # run
    result = smf.logit(formula=formula, data=data).fit()
    # to excel
    for x in x_list:
        p = result.pvalues[x]
        coef = result.params[x].round(2)
        if p >= .05:
            coef = str(coef)
        if p < .05 and p > .01:
            coef = str(coef) + '*'
        if p < .01 and p > .001:
            coef = str(coef) + '**'
        if p < .001:
            coef = str(coef) + '***'
        ws.cell(row=row_n, column=col_n).value = coef
        row_n = row_n + 1
        ws.cell(row=row_n, column=col_n).value = '(' + str(result.bse[x].round(2)) + ')'
        row_n = row_n + 1

    ws.cell(row=row_n, column=col_n).value = result.nobs
    row_n = row_n + 1
    ws.cell(row=row_n, column=col_n).value = result.llf.round(2)
    row_n = row_n + 1
    ws.cell(row=row_n, column=col_n).value = result.prsquared.round(2)
    wb.save(file)

    return result.summary()




def ols_to_excel(data, y, x_list, file, start_row, start_col):
    wb = load_workbook(file)
    ws = wb.active

    col_n = start_col
    row_n = start_row

    # create formula
    i = 1
    x_str = ''
    for x in x_list:
        if i == 1:
            x_str = x_str + x
        if i > 1:
            x_str = x_str + ' + ' + x
        i += 1
    x_str = x_str + '- 1'
    formula = y + ' ~ ' + x_str

    # run
    result = smf.ols(formula=formula, data=data).fit()
    # to excel
    for x in x_list:
        p = result.pvalues[x]
        coef = result.params[x].round(2)
        if p >= .05:
            coef = str(coef)
        if p < .05 and p > .01:
            coef = str(coef) + '*'
        if p < .01 and p > .001:
            coef = str(coef) + '**'
        if p < .001:
            coef = str(coef) + '***'
        ws.cell(row=row_n, column=col_n).value = coef
        row_n = row_n + 1
        ws.cell(row=row_n, column=col_n).value = '(' + str(result.bse[x].round(2)) + ')'
        row_n = row_n + 1

    ws.cell(row=row_n, column=col_n).value = data[y].sum()
    row_n = row_n + 1
    ws.cell(row=row_n, column=col_n).value = result.rsquared.round(2)
    wb.save(file)

    return result.summary()




def var_diff_to_excel(file, df, control_col, diff_col, se_col, pvalue_col, start_col, start_row, change_diff_col = 3):
    wb = load_workbook(file)
    ws = wb.active

    col_n = start_col
    row_n = start_row

    # control
    for ob in df[control_col]:
        ws.cell(row=row_n, column=col_n).value = ob
        row_n = row_n + 2

    # coefficient
    col_n = change_diff_col
    row_n = start_row
    for ob, p in zip(df[diff_col], df[pvalue_col]):
        if p >.05:
            coef = str(ob)
        if p <= .05:
            coef = str(ob) + '*'
        if p < .01:
            coef = str(ob) + '**'
        if p < .001:
            coef = str(ob) + '***'
        ws.cell(row=row_n, column=col_n).value = coef
        row_n = row_n + 2

    # se
    row_n = start_row + 1
    for ob in df[se_col]:
        se = '(' + str(ob) + ')'
        ws.cell(row=row_n, column=col_n).value = se
        row_n = row_n + 2

    wb.save(file)

def just_diff_to_excel(file, df, diff_col, se_col, pvalue_col, start_col, start_row):
    wb = load_workbook(file)
    ws = wb.active

    col_n = start_col

    # coefficient
    row_n = start_row
    for ob, p in zip(df[diff_col], df[pvalue_col]):
        if p >.05:
            coef = str(ob)
        if p <= .05:
            coef = str(ob) + '*'
        if p < .01:
            coef = str(ob) + '**'
        if p < .01:
            coef = str(ob) + '***'
        ws.cell(row=row_n, column=col_n).value = coef
        row_n = row_n + 2

    # se
    row_n = start_row + 1
    for ob in df[se_col]:
        se = '(' + str(ob) + ')'
        ws.cell(row=row_n, column=col_n).value = se
        row_n = row_n + 2

    wb.save(file)

def n_to_excel(file, col, row, n):
    wb = load_workbook(file)
    ws = wb.active


    str_n = 'N = ' + str(n)

    ws.cell(row=row, column=col).value = str_n
    wb.save(file)

